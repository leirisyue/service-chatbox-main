from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, List, Dict
import psycopg2
from psycopg2.extras import RealDictCursor
import google.generativeai as genai
import uuid
import time
import json
from datetime import datetime
from PIL import Image
import os
import re
import pandas as pd
import io
from config import settings
from .textfunc import (format_search_results)

# ========================================
# CONFIGURATION
# ========================================

DB_CONFIG = {
    "dbname": "PRODUCT",
    "user": "postgres",
    "password": "123456",
    "host": "localhost",
    "port": "5432"
}

GEMINI_API_KEY = "AIzaSyCcORIhGV4GbUP9wgSb6FyhzNSw9BUokZ8"
genai.configure(api_key=GEMINI_API_KEY)

app = FastAPI(title="AA Corporation Chatbot API", version="4.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ========================================
# DATABASE HELPERS
# ========================================

def get_db():
    return psycopg2.connect(**DB_CONFIG)

# ========================================
# PYDANTIC MODELS
# ========================================

class ChatMessage(BaseModel):
    session_id: str
    message: str
    context: Optional[Dict] = {}

# ========================================
# GEMINI AI HELPERS
# ========================================

def call_gemini_with_retry(model, prompt, max_retries=3):
    """G·ªçi Gemini v·ªõi retry logic"""
    for attempt in range(max_retries):
        try:
            response = model.generate_content(prompt)
            if response.text:
                return response.text
        except Exception as e:
            if "429" in str(e) or "quota" in str(e).lower():
                wait_time = 5 * (2 ** attempt)
                print(f"‚è≥ Quota exceeded. ƒê·ª£i {wait_time}s...")
                time.sleep(wait_time)
                continue
            print(f"‚ùå L·ªói Gemini: {e}")
            return None
    return None

def generate_embedding(text: str):
    """T·∫°o vector embedding cho text"""
    try:
        result = genai.embed_content(
            model="models/text-embedding-004",
            content=text,
            task_type="retrieval_query"
        )
        return result['embedding']
    except Exception as e:
        print(f"‚ùå L·ªói embedding: {e}")
        return None





# ========================================
# ========================================
# ‚ú® [M·ªöI] HYBRID SEARCH FUNCTIONS
# ========================================

def expand_search_query(user_query: str, params: Dict) -> str:
    """AI m·ªü r·ªông query ng·∫Øn th√†nh m√¥ t·∫£ chi ti·∫øt"""
    model = genai.GenerativeModel("gemini-2.5-flash")
    
    prompt = f"""
Ng∆∞·ªùi d√πng t√¨m: "{user_query}"

T·∫°o m√¥ t·∫£ t√¨m ki·∫øm t·ªëi ∆∞u (2-3 c√¢u ng·∫Øn):
1. LO·∫†I S·∫¢N PH·∫®M (b√†n/gh·∫ø/t·ªß...)
2. V·∫¨T LI·ªÜU C·ª§ TH·ªÇ (g·ªó teak/ƒë√° marble/da b√≤...)
3. V·ªä TR√ç/C√îNG D·ª§NG (nh√† b·∫øp/ph√≤ng kh√°ch/dining/coffee...)

VD: "b√†n g·ªó teak" -> "B√†n l√†m t·ª´ g·ªó teak t·ª± nhi√™n. Dining table ho·∫∑c coffee table ch·∫•t li·ªáu teak wood cao c·∫•p."

Output (ch·ªâ m√¥ t·∫£):
"""
    
    try:
        response = call_gemini_with_retry(model, prompt, max_retries=2)
        if response:
            print(f"‚ú® Expanded: '{user_query}' -> '{response[:80]}...'")
            return response.strip()
    except:
        pass
    return user_query


def extract_product_keywords(query: str) -> list:
    """Tr√≠ch xu·∫•t t·ª´ kh√≥a quan tr·ªçng"""
    materials = ["g·ªó teak", "g·ªó s·ªìi", "g·ªó walnut", "ƒë√° marble", "ƒë√° granite", 
                 "da th·∫≠t", "da b√≤", "v·∫£i linen", "kim lo·∫°i", "teak", "oak", 
                 "walnut", "marble", "granite", "leather"]
    
    contexts = ["nh√† b·∫øp", "ph√≤ng kh√°ch", "ph√≤ng ng·ªß", "vƒÉn ph√≤ng",
                "kitchen", "living room", "dining", "coffee", "bar",
                "b√†n ƒÉn", "b√†n tr√†", "b√†n l√†m vi·ªác"]
    
    shapes = ["tr√≤n", "vu√¥ng", "ch·ªØ nh·∫≠t", "oval", "l-shape", 
              "round", "square", "rectangular"]
    
    types = ["b√†n", "gh·∫ø", "t·ªß", "gi∆∞·ªùng", "sofa", "k·ªá", "ƒë√®n",
             "table", "chair", "cabinet", "bed", "shelf", "lamp"]
    
    query_lower = query.lower()
    keywords = []
    
    for word_list in [materials, contexts, shapes, types]:
        for word in word_list:
            if word in query_lower:
                keywords.append(word)
    
    keywords = list(set(keywords))
    if keywords:
        print(f"üîë Keywords: {keywords}")
    return keywords


def search_products_hybrid(params: Dict):
    """HYBRID: Vector + Keyword Boosting"""
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    # 1. Chu·∫©n b·ªã query
    if params.get("keywords_vector"):
        base = params["keywords_vector"]
    else:
        parts = [params.get("category", ""), params.get("sub_category", ""), 
                params.get("material_primary", "")]
        base = " ".join([p for p in parts if p]) or "n·ªôi th·∫•t"
    
    print(f"\nüîç Query: {base}")
    
    # 2. AI Expansion
    expanded = expand_search_query(base, params)
    
    # 3. Extract keywords
    keywords = extract_product_keywords(expanded)
    
    # 4. Vector
    vector = generate_embedding(expanded)
    if not vector:
        conn.close()
        return {"products": [], "search_method": "failed"}
    
    # 5. SQL Hybrid
    try:
        if keywords:
            conditions = []
            params_list = []
            for kw in keywords:
                conditions.append("(product_name ILIKE %s OR category ILIKE %s OR "
                                "sub_category ILIKE %s OR material_primary ILIKE %s)")
                params_list.extend([f"%{kw}%"] * 4)
            
            boost = f"(CASE WHEN ({' OR '.join(conditions)}) THEN 1 ELSE 0 END)"
        else:
            boost = "0"
            params_list = []
        
        sql = f"""
            SELECT headcode, product_name, category, sub_category, 
                   material_primary, project, project_id,
                   (description_embedding <=> %s::vector) as raw_distance,
                   {boost} as keyword_match
            FROM products
            WHERE description_embedding IS NOT NULL
            ORDER BY (description_embedding <=> %s::vector) - ({boost} * 0.25) ASC
            LIMIT 10
        """
        
        all_params = [vector] + params_list + [vector] + params_list
        cur.execute(sql, all_params)
        results = cur.fetchall()
        
        if results:
            products = [{
                "headcode": r["headcode"],
                "product_name": r["product_name"],
                "category": r.get("category"),
                "sub_category": r.get("sub_category"),
                "material_primary": r.get("material_primary"),
                "project": r.get("project"),
                "project_id": r.get("project_id"),
                "similarity": round(1 - r["raw_distance"], 3),
                "keyword_matched": bool(r.get("keyword_match"))
            } for r in results]
            
            print(f"‚úÖ Found {len(products)} products (Hybrid)")
            conn.close()
            return {
                "products": products,
                "search_method": "hybrid_vector_keyword",
                "expanded_query": expanded
            }
    except Exception as e:
        print(f"‚ùå Hybrid failed: {e}")
    
    conn.close()
    return {"products": [], "search_method": "hybrid_failed"}

# [NEW] AUTO CLASSIFICATION AI
# ========================================

def auto_classify_product(product_name: str, id_sap: str = "") -> Dict:
    """T·ª± ƒë·ªông ph√¢n lo·∫°i s·∫£n ph·∫©m b·∫±ng AI"""
    model = genai.GenerativeModel("gemini-2.5-flash")
    
    prompt = f"""
B·∫°n l√† chuy√™n gia ph√¢n lo·∫°i s·∫£n ph·∫©m n·ªôi th·∫•t cao c·∫•p.

INPUT:
- T√™n s·∫£n ph·∫©m: "{product_name}"
- M√£ SAP: "{id_sap}"

NHI·ªÜM V·ª§: Ph√¢n t√≠ch v√† ph√¢n lo·∫°i s·∫£n ph·∫©m theo 3 ti√™u ch√≠:

1. **category** (Danh m·ª•c ch√≠nh):
   - B√†n (Table)
   - Gh·∫ø (Chair) 
   - Sofa
   - T·ªß (Cabinet)
   - Gi∆∞·ªùng (Bed)
   - ƒê√®n (Lamp)
   - K·ªá (Shelf)
   - B√†n l√†m vi·ªác (Desk)
   - Kh√°c (Other)

2. **sub_category** (Danh m·ª•c ph·ª• - c·ª• th·ªÉ h∆°n):
   VD: "B√†n ƒÉn", "B√†n coffee", "Gh·∫ø bar", "Gh·∫ø ƒÉn", "Sofa g√≥c", "T·ªß qu·∫ßn √°o", "ƒê√®n b√†n", "ƒê√®n tr·∫ßn"...

3. **material_primary** (V·∫≠t li·ªáu ch√≠nh):
   - G·ªó (Wood)
   - Da (Leather)
   - V·∫£i (Fabric)
   - Kim lo·∫°i (Metal)
   - ƒê√° (Stone)
   - K√≠nh (Glass)
   - Nh·ª±a (Plastic)
   - M√¢y tre (Rattan)
   - H·ªón h·ª£p (Mixed)

OUTPUT JSON ONLY (no markdown, no backticks):
{{
  "category": "...",
  "sub_category": "...",
  "material_primary": "..."
}}
"""
    
    response_text = call_gemini_with_retry(model, prompt)
    
    if not response_text:
        return {
            "category": "Ch∆∞a ph√¢n lo·∫°i",
            "sub_category": "Ch∆∞a ph√¢n lo·∫°i", 
            "material_primary": "Ch∆∞a x√°c ƒë·ªãnh"
        }
    
    try:
        clean = response_text.strip()
        if "```json" in clean:
            clean = clean.split("```json")[1].split("```")[0].strip()
        elif "```" in clean:
            clean = clean.split("```")[1].split("```")[0].strip()
        
        result = json.loads(clean)
        return result
    except:
        return {
            "category": "Ch∆∞a ph√¢n lo·∫°i",
            "sub_category": "Ch∆∞a ph√¢n lo·∫°i",
            "material_primary": "Ch∆∞a x√°c ƒë·ªãnh"
        }

def auto_classify_material(material_name: str, id_sap: str = "") -> Dict:
    """T·ª± ƒë·ªông ph√¢n lo·∫°i v·∫≠t li·ªáu b·∫±ng AI"""
    model = genai.GenerativeModel("gemini-2.5-flash")
    
    prompt = f"""
Ph√¢n lo·∫°i nguy√™n v·∫≠t li·ªáu n·ªôi th·∫•t:

T√™n: "{material_name}"
M√£: "{id_sap}"

X√°c ƒë·ªãnh:
1. **material_group**: G·ªó, Da, V·∫£i, ƒê√°, Kim lo·∫°i, K√≠nh, Nh·ª±a, S∆°n, Keo, Ph·ª• ki·ªán, Kh√°c
2. **material_subgroup**: Nh√≥m con c·ª• th·ªÉ (VD: "G·ªó t·ª± nhi√™n", "Da th·∫≠t", "V·∫£i cao c·∫•p"...)

OUTPUT JSON ONLY:
{{
    "material_group": "...",
  "material_subgroup": "..."
}}
"""
    
    response_text = call_gemini_with_retry(model, prompt)
    
    if not response_text:
        return {
            "material_group": "Ch∆∞a ph√¢n lo·∫°i",
            "material_subgroup": "Ch∆∞a ph√¢n lo·∫°i"
        }
    
    try:
        clean = response_text.strip()
        if "```json" in clean:
            clean = clean.split("```json")[1].split("```")[0].strip()
        elif "```" in clean:
            clean = clean.split("```")[1].split("```")[0].strip()
        
        result = json.loads(clean)
        return result
    except:
        return {
            "material_group": "Ch∆∞a ph√¢n lo·∫°i",
            "material_subgroup": "Ch∆∞a ph√¢n lo·∫°i"
        }

# ========================================
# [NEW] CHAT HISTORY
# ========================================

# ========================================
# FIX 1: D√íNG ~430-460
# Thay th·∫ø h√†m save_chat_history
# ========================================

def save_chat_history(session_id: str, user_message: str, bot_response: str, 
                     intent: str, params: Dict, result_count: int,
                     search_type: str = "text",
                     expanded_query: str = None,
                     extracted_keywords: list = None):
    """L∆∞u l·ªãch s·ª≠ chat ƒê·∫¶Y ƒê·ª¶ ƒë·ªÉ h·ªçc - V4.7 FIX"""
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # ‚úÖ QUAN TR·ªåNG: T·∫°o embedding cho query ngay khi l∆∞u
        query_embedding = None
        if user_message:
            query_embedding = generate_embedding(user_message)
        
        sql = """
            INSERT INTO chat_history 
            (session_id, user_message, bot_response, intent, params, result_count,
             search_type, expanded_query, extracted_keywords, query_embedding)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """
        
        cur.execute(sql, (
            session_id, user_message, bot_response, 
            intent, json.dumps(params), result_count,
            search_type,
            expanded_query,
            json.dumps(extracted_keywords) if extracted_keywords else None,
            query_embedding  # ‚úÖ M·ªöI: L∆∞u embedding
        ))
        
        message_id = cur.fetchone()[0]  # ‚úÖ L·∫•y ID message
        
        conn.commit()
        conn.close()
        print(f"üíæ SAVED: msg_id={message_id} | {session_id[:8]}... | {search_type} | {result_count} results")
        
        return message_id  # ‚úÖ Tr·∫£ v·ªÅ ID ƒë·ªÉ UI d√πng
        
    except Exception as e:
        print(f"‚ùå L·ªói save chat history: {e}")
        return None
# ========================================
# HELPER - L·∫§Y GI√Å M·ªöI NH·∫§T
# ========================================

def get_latest_material_price(material_subprice_json: str) -> float:
    """L·∫•y gi√° m·ªõi nh·∫•t t·ª´ JSON l·ªãch s·ª≠ gi√°"""
    if not material_subprice_json:
        print("‚ö†Ô∏è material_subprice_json is NULL/empty")
        return 0.0
    
    try:
        price_history = json.loads(material_subprice_json)
        if not price_history or not isinstance(price_history, list):
            print(f"‚ö†Ô∏è Invalid price_history: {price_history}")
            return 0.0
        
        sorted_prices = sorted(
            price_history, 
            key=lambda x: x.get('date', '1900-01-01'), 
            reverse=True
        )
        result = float(sorted_prices[0].get('price', 0))
        print(f"‚úÖ Parsed price: {result} VNƒê")
        return result
    except:
        print(f"‚ùå Price parse error: {e}")
        return 0.0

# ========================================
# INTENT DETECTION
# ========================================

def get_intent_and_params(user_message: str, context: Dict) -> Dict:
    """AI Router v·ªõi kh·∫£ nƒÉng Reasoning & Soft Clarification"""
    model = genai.GenerativeModel("gemini-2.5-flash")
    
    context_info = ""
    if context.get("current_products"):
        products = context["current_products"]
        context_info = f"\nCONTEXT (User v·ª´a xem): {len(products)} s·∫£n ph·∫©m. SP ƒë·∫ßu ti√™n: {products[0]['headcode']} - {products[0]['product_name']}"
    elif context.get("current_materials"):
        materials = context["current_materials"]
        context_info = f"\nCONTEXT (User v·ª´a xem): {len(materials)} v·∫≠t li·ªáu. VL ƒë·∫ßu ti√™n: {materials[0]['material_name']}"
    
    prompt = f"""
    B·∫°n l√† AI Assistant th√¥ng minh c·ªßa AA Corporation (N·ªôi th·∫•t cao c·∫•p).
    
    INPUT: "{user_message}"
    {context_info}

    NHI·ªÜM V·ª§: Ph√¢n t√≠ch Intent v√† Parameters.
    
    QUY T·∫ÆC SUY LU·∫¨N (LOGIC):
    1. **Intent Detection**: X√°c ƒë·ªãnh user mu·ªën:
       - **search_product**: T√¨m ki·∫øm s·∫£n ph·∫©m (VD: "T√¨m b√†n", "C√≥ b√†n n√†o", "Cho t√¥i xem gh·∫ø")
       - **query_product_materials**: Xem v·∫≠t li·ªáu c·ªßa S·∫¢N PH·∫®M (VD: "V·∫≠t li·ªáu c·ªßa b√†n B001", "Ph√¢n t√≠ch v·∫≠t li·ªáu SP n√†y")
       - **calculate_product_cost**: T√≠nh gi√°/b√°o gi√° S·∫¢N PH·∫®M (VD: "Gi√° b√†n B001", "T√≠nh gi√° s·∫£n ph·∫©m", "B√°o gi√°")
       
       **MATERIAL FLOW:**
       - **search_material**: T√¨m ki·∫øm NGUY√äN V·∫¨T LI·ªÜU (VD: "T√¨m g·ªó s·ªìi", "C√≥ lo·∫°i da n√†o", "ƒê√° marble", "V·∫≠t li·ªáu l√†m b√†n")
       - **query_material_detail**: Xem chi ti·∫øt V·∫¨T LI·ªÜU + s·∫£n ph·∫©m s·ª≠ d·ª•ng (VD: "Chi ti·∫øt g·ªó s·ªìi", "Xem v·∫≠t li·ªáu n√†y d√πng ·ªü ƒë√¢u")
       - **list_material_groups**: Li·ªát k√™ nh√≥m v·∫≠t li·ªáu (VD: "C√°c lo·∫°i g·ªó", "Danh s√°ch ƒë√°")

       ----------------------------------------------------------------
       **[NEW] CROSS-TABLE INTENTS (B·ªî SUNG ‚Äì KH√îNG THAY ƒê·ªîI LOGIC C≈®):**
       - **search_product_by_material**: T√¨m s·∫£n ph·∫©m L√ÄM T·ª™ v·∫≠t li·ªáu c·ª• th·ªÉ
         V√≠ d·ª•: "T√¨m b√†n l√†m t·ª´ ƒë√° marble", "T·ªß g·ªó teak", "Gh·∫ø da th·∫≠t"
       
       - **search_material_for_product**: T√¨m v·∫≠t li·ªáu ƒê·ªÇ L√ÄM s·∫£n ph·∫©m c·ª• th·ªÉ
         V√≠ d·ª•: "V·∫≠t li·ªáu l√†m b√†n tr√≤n", "Nguy√™n li·ªáu gh·∫ø sofa", "ƒê√° l√†m b√†n"

       **PH√ÇN BI·ªÜT R√ï (∆ØU TI√äN TU√ÇN TH·ª¶):**
       - "T√¨m b√†n g·ªó" ‚Üí search_product
       - "T√¨m b√†n L√ÄM T·ª™ g·ªó teak" ‚Üí search_product_by_material
       - "T√¨m g·ªó" ‚Üí search_material
       - "T√¨m v·∫≠t li·ªáu ƒê·ªÇ L√ÄM b√†n" ‚Üí search_material_for_product
       ----------------------------------------------------------------
       
       - **greeting**: Ch√†o h·ªèi (VD: "Xin ch√†o", "Hello", "Hi")
       - **unknown**: Kh√¥ng r√µ √Ω ƒë·ªãnh
    
    2. **Entity Type Detection**: 
       - Ph√¢n bi·ªát: User ƒëang n√≥i v·ªÅ S·∫¢N PH·∫®M hay V·∫¨T LI·ªÜU?
       - Keyword: "s·∫£n ph·∫©m", "b√†n", "gh·∫ø", "sofa" ‚Üí PRODUCT
       - Keyword: "v·∫≠t li·ªáu", "nguy√™n li·ªáu", "g·ªó", "da", "ƒë√°", "v·∫£i" ‚Üí MATERIAL
       - "gi√°" + context s·∫£n ph·∫©m ‚Üí calculate_product_cost
       - "gi√°" + context v·∫≠t li·ªáu ‚Üí query_material_detail
    
    3. **Broad Query Detection**: 
       - N·∫øu User ch·ªâ n√≥i danh m·ª•c l·ªõn (VD: "T√¨m b√†n", "Gh·∫ø", "ƒê√®n", "T√¨m g·ªó") m√† KH√îNG c√≥ t√≠nh ch·∫•t c·ª• th·ªÉ:
         -> Set `is_broad_query`: true
         -> T·∫°o `follow_up_question`: M·ªôt c√¢u h·ªèi ng·∫Øn g·ª£i √Ω user thu h·∫πp ph·∫°m vi
       - N·∫øu User ƒë√£ c·ª• th·ªÉ (VD: "B√†n ƒÉn tr√≤n", "Gh·∫ø g·ªó s·ªìi", "ƒê√° marble tr·∫Øng"):
         -> Set `is_broad_query`: false
         -> `follow_up_question`: null
    
    4. **Parameter Extraction**:
       **For PRODUCTS:**
       - `category`: Danh m·ª•c s·∫£n ph·∫©m
       - `sub_category`: Danh m·ª•c ph·ª•
       - `material_primary`: V·∫≠t li·ªáu ch√≠nh
       - `keywords_vector`: M√¥ t·∫£ ƒë·∫ßy ƒë·ªß ƒë·ªÉ search vector
       - `headcode`: M√£ s·∫£n ph·∫©m (n·∫øu c√≥ trong INPUT ho·∫∑c Context)
       
       **For MATERIALS:**
       - `material_name`: T√™n v·∫≠t li·ªáu (VD: "g·ªó s·ªìi", "da th·∫≠t")
       - `material_group`: Nh√≥m v·∫≠t li·ªáu (VD: "G·ªó", "Da", "ƒê√°", "V·∫£i")
       - `material_subgroup`: Nh√≥m con
       - `keywords_vector`: M√¥ t·∫£ ƒë·∫∑c t√≠nh ƒë·ªÉ search (VD: "g·ªó l√†m b√†n ƒÉn cao c·∫•p m√†u n√¢u")
       - `id_sap`: M√£ v·∫≠t li·ªáu SAP (n·∫øu c√≥)
       - `usage_context`: Ng·ªØ c·∫£nh s·ª≠ d·ª•ng (VD: "l√†m b√†n", "b·ªçc gh·∫ø")
    
    5. **Context Awareness**:
       - N·∫øu User d√πng t·ª´ ƒë·∫°i t·ª´ ("c√°i n√†y", "n√≥", "s·∫£n ph·∫©m ƒë√≥", "v·∫≠t li·ªáu n√†y"), h√£y l·∫•y t·ª´ Context
       - N·∫øu User h·ªèi v·ªÅ gi√°/v·∫≠t li·ªáu m√† kh√¥ng n√≥i r√µ, ∆∞u ti√™n l·∫•y item ƒë·∫ßu ti√™n trong Context

    OUTPUT FORMAT (JSON ONLY - no markdown backticks):
    {{
      "intent": "search_product|search_product_by_material|search_material_for_product|query_product_materials|calculate_product_cost|search_material|query_material_detail|list_material_groups|greeting|unknown",
      "entity_type": "product|material|unknown",
      "params": {{
        "category": "String ho·∫∑c null",
        "sub_category": "String ho·∫∑c null",
        "material_primary": "String ho·∫∑c null",
        "material_name": "String ho·∫∑c null",
        "material_group": "String ho·∫∑c null",
        "material_subgroup": "String ho·∫∑c null",
        "keywords_vector": "T·ª´ kh√≥a m√¥ t·∫£ ƒë·∫ßy ƒë·ªß",
        "headcode": "String ho·∫∑c null",
        "id_sap": "String ho·∫∑c null",
        "usage_context": "String ho·∫∑c null"
      }},
      "is_broad_query": boolean,
      "follow_up_question": "String ho·∫∑c null",
      "suggested_actions": ["String 1", "String 2"]
    }}
    """
    
    response_text = call_gemini_with_retry(model, prompt)
    if not response_text:
        return {"intent": "error"}
    
    try:
        clean_text = response_text.strip()
        
        if "```json" in clean_text:
            clean_text = clean_text.split("```json")[1].split("```")[0].strip()
        elif "```" in clean_text:
            clean_text = clean_text.split("```")[1].split("```")[0].strip()
        
        result = json.loads(clean_text)
        
        if result["intent"] in ["calculate_product_cost", "query_product_materials"]:
            if not result["params"].get("headcode"):
                match = re.search(r'\b([A-Z0-9]+-?[A-Z0-9]+)\b', user_message.upper())
                if match:
                    result["params"]["headcode"] = match.group(1)
        
        return result
        
    except json.JSONDecodeError as e:
        print(f"JSON Parse Error: {e} - Raw: {response_text}")
        return {"intent": "error", "raw": response_text}
    except Exception as e:
        print(f"Parse Error: {e}")
        return {"intent": "error", "raw": response_text}

# ========================================
# [NEW] CROSS-TABLE SEARCH FUNCTIONS
# ========================================

def search_products_by_material(material_query: str, params: Dict):
    """
    üîç T√åM S·∫¢N PH·∫®M ƒê∆Ø·ª¢C L√ÄM T·ª™ V·∫¨T LI·ªÜU C·ª§ TH·ªÇ
    V√≠ d·ª•: "T√¨m b√†n l√†m t·ª´ ƒë√° marble", "T·ªß g·ªó teak"
    
    Logic: 
    1. T√¨m materials ph√π h·ª£p v·ªõi query (vector search)
    2. JOIN product_materials ƒë·ªÉ l·∫•y products s·ª≠ d·ª•ng material ƒë√≥
    3. Rank products theo ƒë·ªô ph√π h·ª£p
    """
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    print(f"üîó Cross-table search: Products made from '{material_query}'")
    
    # B∆∞·ªõc 1: T√¨m v·∫≠t li·ªáu ph√π h·ª£p
    material_vector = generate_embedding(material_query)
    
    if not material_vector:
        conn.close()
        return {"products": [], "search_method": "failed"}
    
    try:
        # T√¨m top materials ph√π h·ª£p
        cur.execute("""
            SELECT 
                id_sap, 
                material_name,
                material_group,
                (description_embedding <=> %s::vector) as distance
            FROM materials
            WHERE description_embedding IS NOT NULL
            ORDER BY distance ASC
            LIMIT 5
        """, [material_vector])
        
        matched_materials = cur.fetchall()
        
        if not matched_materials:
            conn.close()
            return {"products": [], "search_method": "no_materials_found"}
        
        material_ids = [m['id_sap'] for m in matched_materials]
        material_names = [m['material_name'] for m in matched_materials]
        
        print(f"‚úÖ Found {len(material_ids)} matching materials: {material_names[:3]}")
        
        # B∆∞·ªõc 2: T√¨m products s·ª≠ d·ª•ng materials n√†y
        # K·∫øt h·ª£p filter category n·∫øu c√≥
        category_filter = ""
        filter_params = []
        
        if params.get("category"):
            category_filter = "AND p.category ILIKE %s"
            filter_params.append(f"%{params['category']}%")
        
        sql = f"""
            SELECT 
                p.headcode,
                p.product_name,
                p.category,
                p.sub_category,
                p.material_primary,
                p.project,
                p.project_id,
                m.material_name,
                m.id_sap as material_id,
                pm.quantity,
                COUNT(*) OVER (PARTITION BY p.headcode) as material_match_count
            FROM products p
            INNER JOIN product_materials pm ON p.headcode = pm.product_headcode
            INNER JOIN materials m ON pm.material_id_sap = m.id_sap
            WHERE m.id_sap = ANY(%s)
            {category_filter}
            ORDER BY material_match_count DESC, p.product_name ASC
            LIMIT 20
        """
        
        cur.execute(sql, [material_ids] + filter_params)
        results = cur.fetchall()
        
        conn.close()
        
        if not results:
            return {
                "products": [],
                "search_method": "cross_table_no_products",
                "matched_materials": material_names
            }
        
        # Group products (v√¨ 1 product c√≥ th·ªÉ d√πng nhi·ªÅu materials)
        products_dict = {}
        for row in results:
            headcode = row['headcode']
            if headcode not in products_dict:
                products_dict[headcode] = {
                    "headcode": headcode,
                    "product_name": row['product_name'],
                    "category": row['category'],
                    "sub_category": row['sub_category'],
                    "material_primary": row['material_primary'],
                    "project": row['project'],
                    "project_id": row['project_id'],
                    "matched_materials": [],
                    "relevance_score": 0
                }
            
            products_dict[headcode]["matched_materials"].append({
                "name": row['material_name'],
                "id": row['material_id'],
                "quantity": row['quantity']
            })
            products_dict[headcode]["relevance_score"] += 1
        
        products_list = sorted(
            products_dict.values(),
            key=lambda x: x['relevance_score'],
            reverse=True
        )
        
        print(f"‚úÖ Found {len(products_list)} products using these materials")
        
        return {
            "products": products_list[:10],
            "search_method": "cross_table_material_to_product",
            "matched_materials": material_names,
            "explanation": f"T√¨m th·∫•y s·∫£n ph·∫©m s·ª≠ d·ª•ng: {', '.join(material_names[:3])}"
        }
        
    except Exception as e:
        print(f"‚ùå Cross-table search failed: {e}")
        conn.close()
        return {"products": [], "search_method": "cross_table_error"}



def search_materials_for_product(product_query: str, params: Dict):
    """
    üîç T√åM V·∫¨T LI·ªÜU ƒê·ªÇ L√ÄM S·∫¢N PH·∫®M C·ª§ TH·ªÇ
    V√≠ d·ª•: "V·∫≠t li·ªáu l√†m b√†n tr√≤n", "Nguy√™n li·ªáu gh·∫ø sofa"
    
    Logic:
    1. T√¨m products ph√π h·ª£p v·ªõi query
    2. JOIN product_materials ƒë·ªÉ l·∫•y materials ƒë∆∞·ª£c d√πng
    3. Aggregate + rank theo t·∫ßn su·∫•t
    """
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    print(f"üîó Cross-table search: Materials for '{product_query}'")
    
    # B∆∞·ªõc 1: T√¨m products ph√π h·ª£p
    product_vector = generate_embedding(product_query)
    
    if not product_vector:
        conn.close()
        return {"materials": [], "search_method": "failed"}
    
    try:
        cur.execute("""
            SELECT 
                headcode,
                product_name,
                category,
                (description_embedding <=> %s::vector) as distance
            FROM products
            WHERE description_embedding IS NOT NULL
            ORDER BY distance ASC
            LIMIT 10
        """, [product_vector])
        
        matched_products = cur.fetchall()
        
        if not matched_products:
            conn.close()
            return {"materials": [], "search_method": "no_products_found"}
        
        product_headcodes = [p['headcode'] for p in matched_products]
        product_names = [p['product_name'] for p in matched_products]
        
        print(f"‚úÖ Found {len(product_headcodes)} matching products: {product_names[:3]}")
        
        # B∆∞·ªõc 2: L·∫•y materials ƒë∆∞·ª£c d√πng trong products n√†y
        material_filter = ""
        filter_params = []
        
        if params.get("material_group"):
            material_filter = "AND m.material_group ILIKE %s"
            filter_params.append(f"%{params['material_group']}%")
        
        sql = f"""
            SELECT 
                m.id_sap,
                m.material_name,
                m.material_group,
                m.material_subgroup,
                m.material_subprice,
                m.unit,
                m.image_url,
                COUNT(DISTINCT pm.product_headcode) as usage_count,
                SUM(pm.quantity) as total_quantity,
                array_agg(DISTINCT p.product_name) as used_in_products
            FROM materials m
            INNER JOIN product_materials pm ON m.id_sap = pm.material_id_sap
            INNER JOIN products p ON pm.product_headcode = p.headcode
            WHERE p.headcode = ANY(%s)
            {material_filter}
            GROUP BY m.id_sap, m.material_name, m.material_group, 
                     m.material_subgroup, m.material_subprice, m.unit, m.image_url
            ORDER BY usage_count DESC, m.material_name ASC
            LIMIT 15
        """
        
        cur.execute(sql, [product_headcodes] + filter_params)
        results = cur.fetchall()
        
        conn.close()
        
        if not results:
            return {
                "materials": [],
                "search_method": "cross_table_no_materials",
                "matched_products": product_names
            }
        
        materials_with_context = []
        for mat in results:
            mat_dict = dict(mat)
            mat_dict['price'] = get_latest_material_price(mat['material_subprice'])
            mat_dict['used_in_products_list'] = mat['used_in_products'][:5]  # Top 5
            mat_dict['relevance_score'] = mat['usage_count'] / max_usage if max_usage > 0 else 0.5
            materials_with_context.append(mat_dict)
        
        print(f"‚úÖ Found {len(materials_with_context)} materials used in these products")
        
        return {
            "materials": materials_with_context,
            "search_method": "cross_table_product_to_material",
            "matched_products": product_names[:5],
            "explanation": f"V·∫≠t li·ªáu th∆∞·ªùng d√πng cho: {', '.join(product_names[:3])}"
        }
        
    except Exception as e:
        print(f"‚ùå Cross-table materials search failed: {e}")
        conn.close()
        return {"materials": [], "search_method": "cross_table_error"}


# ========================================
# [NEW] USER FEEDBACK LEARNING SYSTEM
# ========================================


# ========================================
# THAY TH·∫æ h√†m save_user_feedback (d√≤ng ~615)
# ========================================

def save_user_feedback(session_id: str, query: str, selected_items: list, 
                       rejected_items: list, search_type: str):
    """
    üíæ V5.1 - L∆∞u feedback V√Ä embedding cho query
    """
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # ‚úÖ T·∫†O EMBEDDING CHO QUERY NGAY KHI L∆ØU
        query_embedding = generate_embedding(query)
        
        if not query_embedding:
            print("‚ö†Ô∏è Kh√¥ng t·∫°o ƒë∆∞·ª£c embedding, v·∫´n l∆∞u feedback")
        
        sql = """
            INSERT INTO user_feedback 
            (session_id, query, selected_items, rejected_items, search_type, query_embedding)
            VALUES (%s, %s, %s, %s, %s, %s)
            RETURNING id
        """
        
        cur.execute(sql, (
            session_id,
            query,
            json.dumps(selected_items),
            json.dumps(rejected_items),
            search_type,
            query_embedding  # ‚úÖ L∆ØU EMBEDDING
        ))
        
        feedback_id = cur.fetchone()[0]
        
        conn.commit()
        conn.close()
        
        print(f"üíæ Feedback saved: {len(selected_items)} selected, {len(rejected_items)} rejected")
        print(f"   ‚Üí Feedback ID: {feedback_id}")
        print(f"   ‚Üí Embedding: {'‚úÖ OK' if query_embedding else '‚ùå NULL'}")
        
        return True
        
    except Exception as e:
        print(f"‚ùå Failed to save feedback: {e}")
        import traceback
        traceback.print_exc()
        return False






# ======================================
# THAY TH·∫æ h√†m get_feedback_boost_for_query (d√≤ng ~900)
# ========================================

def get_feedback_boost_for_query(query: str, search_type: str, similarity_threshold: float = None) -> Dict:
    """
    üìä V5.0 - Vector-based feedback matching
    T√¨m feedback t·ª´ c√°c query T∆Ø∆†NG T·ª∞ (kh√¥ng c·∫ßn tr√πng 100%)
    
    Args:
        query: C√¢u h·ªèi hi·ªán t·∫°i
        search_type: "product" ho·∫∑c "material"
        similarity_threshold: Ng∆∞·ª°ng ƒë·ªô t∆∞∆°ng t·ª± (m·∫∑c ƒë·ªãnh t·ª´ config)
    
    Returns:
        Dict[item_id, feedback_score]
    """
    from config import settings
    if similarity_threshold is None:
        similarity_threshold = settings.SIMILARITY_THRESHOLD_HIGH
    
    try:
        # 1. T·∫°o embedding cho query hi·ªán t·∫°i
        query_vector = generate_embedding(query)
        
        if not query_vector:
            print("‚ùå Kh√¥ng t·∫°o ƒë∆∞·ª£c embedding cho query")
            return {}
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # 2. T√¨m c√°c feedback c√≥ query_embedding t∆∞∆°ng t·ª± (cosine similarity)
        cur.execute("""
            SELECT 
                query,
                selected_items,
                (1 - (query_embedding <=> %s::vector)) as similarity
            FROM user_feedback
            WHERE search_type = %s
              AND query_embedding IS NOT NULL
              AND (1 - (query_embedding <=> %s::vector)) >= %s
            ORDER BY similarity DESC
            LIMIT 20
        """, (query_vector, search_type, query_vector, similarity_threshold))
        
        similar_feedbacks = cur.fetchall()
        conn.close()
        
        if not similar_feedbacks:
            print(f"‚ÑπÔ∏è Kh√¥ng c√≥ feedback t∆∞∆°ng t·ª± (threshold={similarity_threshold})")
            return {}
        
        # 3. T√≠nh ƒëi·ªÉm cho t·ª´ng item (weighted by similarity)
        item_scores = {}
        
        print(f"\n{'='*60}")
        print(f"üìä FEEDBACK BOOST: T√¨m th·∫•y {len(similar_feedbacks)} query t∆∞∆°ng t·ª±")
        print(f"{'='*60}\n")
        
        for fb in similar_feedbacks:
            sim = fb['similarity']
            
            try:
                # ‚úÖ FIX: Ki·ªÉm tra type tr∆∞·ªõc khi parse
                selected_items = fb['selected_items']
                
                # N·∫øu l√† string JSON ‚Üí parse
                if isinstance(selected_items, str):
                    selected = json.loads(selected_items)
                # N·∫øu ƒë√£ l√† list ‚Üí d√πng lu√¥n
                elif isinstance(selected_items, list):
                    selected = selected_items
                else:
                    print(f"‚ö†Ô∏è Unknown type for selected_items: {type(selected_items)}")
                    continue
                
                print(f"‚úÖ Query: '{fb['query'][:50]}...' (sim={sim:.2f})")
                print(f"   ‚Üí Selected: {selected[:3]}")
                
                for item_id in selected:
                    # ƒêi·ªÉm = similarity * 1 (c√≥ th·ªÉ thay b·∫±ng decay theo th·ªùi gian)
                    item_scores[item_id] = item_scores.get(item_id, 0) + sim
                    
            except Exception as e:
                print(f"‚ö†Ô∏è Skip feedback: {e}")
                continue
        
        if item_scores:
            print(f"\nüìà K·∫øt qu·∫£:")
            for item_id, score in sorted(item_scores.items(), key=lambda x: x[1], reverse=True)[:5]:
                print(f"   {item_id}: {score:.2f} ƒëi·ªÉm")
        else:
            print("‚ÑπÔ∏è Kh√¥ng c√≥ item n√†o ƒë∆∞·ª£c boost")
            
        print(f"{'='*60}\n")
        
        return item_scores
        
    except Exception as e:
        print(f"‚ùå Failed to get feedback boost: {e}")
        import traceback
        traceback.print_exc()
        return {}
# ========================================
# THAY TH·∫æ h√†m rerank_with_feedback (d√≤ng ~570)
# Th√™m LOG chi ti·∫øt
# ========================================

def rerank_with_feedback(items: list, feedback_scores: Dict, 
                         id_key: str = "headcode", boost_weight: float = 0.5):  # ‚Üê ‚úÖ TƒÇNG t·ª´ 0.3 ‚Üí 0.5
    """
    üéØ V5.6 - Boost weight tƒÉng l√™n 0.5 ƒë·ªÉ feedback c√≥ t√°c ƒë·ªông m·∫°nh h∆°n
    """
    if not feedback_scores:
        print("‚ö†Ô∏è Kh√¥ng c√≥ feedback scores ƒë·ªÉ rerank")
        return items
    
    max_feedback = max(feedback_scores.values()) if feedback_scores else 1
    
    print(f"\n{'='*60}")
    print(f"üéØ RERANKING: {len(items)} items | Boost weight: {boost_weight}")
    print(f"üìä Feedback history: {len(feedback_scores)} items c√≥ ƒëi·ªÉm")
    print(f"{'='*60}\n")
    
    boosted_items = []
    unchanged_items = []
    
    for item in items:
        item_id = item.get(id_key)
        feedback_count = feedback_scores.get(item_id, 0)
        
        # Normalize feedback score 0-1
        feedback_boost = (feedback_count / max_feedback) if max_feedback > 0 else 0
        
        # ‚úÖ QUAN TR·ªåNG: D√πng 'similarity' (ƒë√£ ƒë∆∞·ª£c set = personalized_score)
        current_score = item.get('similarity', item.get('relevance_score', 0.5))
        
        # ‚úÖ C√¥ng th·ª©c m·ªõi: Boost weight cao h∆°n (0.5 thay v√¨ 0.3)
        new_score = (1 - boost_weight) * current_score + boost_weight * feedback_boost
        
        item['final_score'] = float(new_score)
        item['feedback_boost'] = float(feedback_boost)
        item['feedback_count'] = float(feedback_count)
        item['original_score'] = float(current_score)
        
        if feedback_count > 0:
            boosted_items.append(item)
            print(f"‚úÖ BOOSTED: {item_id[:20]:20} | "
                  f"Original: {current_score:.3f} ‚Üí "
                  f"Final: {new_score:.3f} | "
                  f"Feedback: {feedback_count:.2f} l·∫ßn")
        else:
            unchanged_items.append(item)
    
    print(f"\nüìà K·∫øt qu·∫£:")
    print(f"   - {len(boosted_items)} items ƒë∆∞·ª£c boost")
    print(f"   - {len(unchanged_items)} items kh√¥ng ƒë·ªïi")
    print(f"{'='*60}\n")
    
    return items  # Kh√¥ng sort ·ªü ƒë√¢y, ƒë·ªÉ search_products() sort sau
# ========================================
# TH√äM V√ÄO chatbot_api.py SAU H√ÄM rerank_with_feedback
# D√≤ng ~620
# ========================================

# ========================================
# S·ª¨A trong apply_feedback_to_search (d√≤ng ~720)
# ========================================

def apply_feedback_to_search(items: list, query: str, search_type: str, 
                             id_key: str = "headcode") -> list:
    """
    üéØ V5.6 - L∆∞u original_rank TR∆Ø·ªöC khi rerank
    """
    if not items:
        return items
    
    # ‚úÖ L∆ØU ORIGINAL RANK (d·ª±a tr√™n personalized_score)
    for idx, item in enumerate(items):
        item['original_rank'] = idx + 1
    
    # Get feedback scores
    feedback_scores = get_feedback_boost_for_query(
        query, 
        search_type,
        similarity_threshold=settings.SIMILARITY_THRESHOLD_VERY_HIGH
    )
    
    if not feedback_scores:
        print("‚ÑπÔ∏è Kh√¥ng c√≥ feedback history ph√π h·ª£p")
        for item in items:
            item['has_feedback'] = False
            item['feedback_count'] = 0
            item['final_rank'] = items.index(item) + 1
            item['final_score'] = item.get('similarity', 0.5)
        return items
    
    print(f"\nüéØ Step 2: Feedback Ranking for {len(items)} items...")
    
    # Apply reranking
    reranked_items = rerank_with_feedback(
        items, 
        feedback_scores, 
        id_key=id_key, 
        boost_weight=0.5  # ‚úÖ Boost weight cao
    )
    
    # ‚úÖ SORT theo final_score (search_products s·∫Ω sort l·∫°i l·∫ßn cu·ªëi)
    reranked_items.sort(key=lambda x: x.get('final_score', 0), reverse=True)
    
    # Update final rank
    for idx, item in enumerate(reranked_items):
        item['final_rank'] = idx + 1
        item['has_feedback'] = item.get('feedback_count', 0) > 0
    
    print(f"‚úÖ Feedback Ranking done\n")
    return reranked_items
# ========================================
# HO·∫∂C L√ÄM THRESHOLD ƒê·ªòNG (t√πy ch·ªçn)
# ========================================

def get_adaptive_threshold(query: str) -> float:
    """
    T·ª± ƒë·ªông ƒëi·ªÅu ch·ªânh threshold:
    - Query d√†i, c·ª• th·ªÉ ‚Üí threshold th·∫•p (0.75)
    - Query ng·∫Øn, chung chung ‚Üí threshold cao (0.90)
    """
    words = query.split()
    
    if len(words) >= 8:
        return 0.75  # Query d√†i ‚Üí d·ªÖ d√£i h∆°n
    elif len(words) >= 5:
        return 0.82
    else:
        return 0.90  # Query ng·∫Øn ‚Üí nghi√™m ng·∫∑t h∆°n
    
# D√πng trong apply_feedback_to_search:
# threshold = get_adaptive_threshold(query)
# feedback_scores = get_feedback_boost_for_query(query, search_type, threshold)


def get_ranking_summary(items: list) -> dict:
    """
    üìä T·∫°o summary v·ªÅ ranking ƒë·ªÉ hi·ªÉn th·ªã trong UI
    
    Returns:
        {
            "total_items": 10,
            "boosted_items": 3,
            "max_boost": 5,
            "ranking_changes": [
                {"id": "B001", "from": 5, "to": 1},
                ...
            ]
        }
    """
    if not items:
        return {
            "total_items": 0,
            "boosted_items": 0,
            "ranking_applied": False
        }
    
    boosted = [i for i in items if i.get('feedback_count', 0) > 0]
    
    changes = []
    for item in items:
        orig = item.get('original_rank')
        final = item.get('final_rank')
        
        if orig and final and orig != final:
            changes.append({
                "id": item.get('headcode') or item.get('id_sap'),
                "name": (item.get('product_name') or item.get('material_name', ''))[:30],
                "from_rank": orig,
                "to_rank": final,
                "boost": orig - final  # Positive = moved up
            })
    
    # Sort by biggest boost first
    changes.sort(key=lambda x: x['boost'], reverse=True)
    
    return {
        "total_items": len(items),
        "boosted_items": len(boosted),
        "ranking_applied": len(boosted) > 0,
        "max_feedback_count": max([i.get('feedback_count', 0) for i in items]),
        "ranking_changes": changes[:5]  # Top 5 changes
    }
# ========================================
# PRODUCT FUNCTIONS
# ========================================

def search_products(params: Dict, session_id: str = None):
    """üîç V5.9 - FIXED: Ch·ªâ 1 l·ªõp ranking, kh√¥ng overlap"""
    
    try:
        result = search_products_hybrid(params)
        if result.get("products"):
            products = result["products"]
            
            # ========== STEP 1: BASE SCORES ==========
            for product in products:
                product['base_score'] = float(product.get('similarity', 0.5))
            
            # ========== STEP 2: PERSONALIZATION ==========
            # ‚úÖ CH·ªà √°p d·ª•ng n·∫øu c√≥ session_id V√Ä user c√≥ history
            has_personalization = False
            
            if session_id:
                print(f"\nüéØ Personalization for {session_id[:8]}...")
                
                # ‚úÖ CHECK tr∆∞·ªõc xem user c√≥ history kh√¥ng
                conn = get_db()
                cur = conn.cursor()
                cur.execute("""
                    SELECT COUNT(*) FROM user_preferences 
                    WHERE session_id = %s
                """, (session_id,))
                history_count = cur.fetchone()[0]
                conn.close()
                
                if history_count > 0:
                    has_personalization = True
                    print(f"   ‚úÖ Found {history_count} interactions")
                    
                    for product in products:
                        conn = get_db()
                        cur = conn.cursor(cursor_factory=RealDictCursor)
                        
                        cur.execute("""
                            SELECT description_embedding 
                            FROM products 
                            WHERE headcode = %s AND description_embedding IS NOT NULL
                        """, (product['headcode'],))
                        
                        vec_result = cur.fetchone()
                        conn.close()
                        
                        if vec_result and vec_result['description_embedding']:
                            personal_score = calculate_personalized_score(
                                vec_result['description_embedding'],
                                session_id
                            )
                            product['personal_score'] = float(personal_score)
                        else:
                            product['personal_score'] = 0.5
                else:
                    print(f"   ‚ÑπÔ∏è No history - Skip personalization")
            
            # ‚úÖ N·∫øu kh√¥ng c√≥ personalization ‚Üí set neutral 0.5
            if not has_personalization:
                for product in products:
                    product['personal_score'] = 0.5
            
            print(f"‚úÖ Personalization done\n")
            
            # ========== STEP 3: FEEDBACK SCORES ==========
            print(f"üéØ Feedback Scoring...")
            
            feedback_dict = get_feedback_boost_for_query(
                params.get("keywords_vector", ""),
                search_type="product",
                similarity_threshold=settings.SIMILARITY_THRESHOLD_VERY_HIGH
            )
            
            max_feedback = max(feedback_dict.values()) if feedback_dict else 1.0
            
            for product in products:
                headcode = product.get('headcode')
                raw_feedback = feedback_dict.get(headcode, 0)
                
                product['feedback_score'] = float(raw_feedback / max_feedback) if max_feedback > 0 else 0.0
                product['feedback_count'] = float(raw_feedback)
            
            print(f"‚úÖ Feedback Scoring done\n")
            
            # ========== STEP 4: WEIGHTED SUM ==========
            print(f"üéØ Final Ranking (Weighted Sum)...")
            
            # ‚úÖ ADAPTIVE WEIGHTS
            if has_personalization:
                # User c√≥ history ‚Üí ∆∞u ti√™n personalization
                W_BASE = 0.3
                W_PERSONAL = 0.5
                W_FEEDBACK = 0.2
            else:
                # User m·ªõi ‚Üí ∆∞u ti√™n base + social proof
                W_BASE = 0.6
                W_PERSONAL = 0.0  # ‚ùå KH√îNG d√πng personal_score
                W_FEEDBACK = 0.4
            
            for idx, product in enumerate(products):
                base = product.get('base_score', 0.5)
                personal = product.get('personal_score', 0.5)
                feedback = product.get('feedback_score', 0.0)
                
                # ‚úÖ Ch·ªâ t√≠nh personal n·∫øu has_personalization
                if has_personalization:
                    final_score = (W_BASE * base) + (W_PERSONAL * personal) + (W_FEEDBACK * feedback)
                else:
                    final_score = (W_BASE * base) + (W_FEEDBACK * feedback)
                
                product['final_score'] = float(final_score)
                product['original_rank'] = idx + 1
                
                print(f"  {product['headcode']}: "
                      f"base={base:.3f} | pers={personal:.3f} | fb={feedback:.3f} "
                      f"‚Üí final={final_score:.3f}")
            
            # ========== STEP 5: SORT FINAL ==========
            products.sort(key=lambda x: x.get('final_score', 0), reverse=True)
            
            for idx, product in enumerate(products):
                product['final_rank'] = idx + 1
                
                if product.get('feedback_count', 0) > 0:
                    product['has_feedback'] = True
            
            print(f"‚úÖ Final Ranking complete\n")
            
            result["products"] = products
            result["ranking_summary"] = get_ranking_summary(products)
            result["can_provide_feedback"] = True
            
            return result
            
    except Exception as e:
        print(f"‚ö†Ô∏è TIER 1 failed: {e}")
        import traceback
        traceback.print_exc()
    
    # TIER 2 & 3: Gi·ªØ nguy√™n code c≈© (Fallback)
    # ...
    
    # TIER 2 & 3: Gi·ªØ nguy√™n...
    # (code c≈© kh√¥ng ƒë·ªïi)
    
    # TIER 2 & 3: GI·ªÆ NGUY√äN CODE C≈® (Fallback)
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    if params.get("keywords_vector"):
        query_text = params["keywords_vector"]
    else:
        query_parts = []
        if params.get("category"): query_parts.append(params["category"])
        if params.get("sub_category"): query_parts.append(params["sub_category"])
        if params.get("material_primary"): query_parts.append(params["material_primary"])
        query_text = " ".join(query_parts) if query_parts else "n·ªôi th·∫•t"

    query_vector = generate_embedding(query_text)
    
    if not query_vector:
        conn.close()
        return search_products_keyword_only(params)
    
    # TIER 2: Pure Vector
    try:
        sql = """
            SELECT headcode, product_name, category, sub_category, 
                   material_primary, project, project_id,
                   (description_embedding <=> %s::vector) as distance
            FROM products
            WHERE description_embedding IS NOT NULL
            ORDER BY distance ASC
            LIMIT 10
        """
        
        cur.execute(sql, [query_vector])
        results = cur.fetchall()
        
        if results:
            print(f"‚úÖ TIER 2: {len(results)} products")
            products = format_search_results(results[:8])
            conn.close()
            return {"products": products, "search_method": "vector_no_filter"}
    except Exception as e:
        print(f"‚ö†Ô∏è TIER 2 failed: {e}")
    
    # TIER 3: Keyword
    conn.close()
    return search_products_keyword_only(params)

def search_products_keyword_only(params: Dict):
    """TIER 3: Fallback keyword search"""
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    conditions = []
    values = []
    
    if params.get("category"):
        cat = params['category']
        conditions.append("(category ILIKE %s OR sub_category ILIKE %s OR product_name ILIKE %s)")
        values.extend([f"%{cat}%", f"%{cat}%", f"%{cat}%"])
    
    if params.get("material_primary"):
        mat = params['material_primary']
        conditions.append("(material_primary ILIKE %s OR product_name ILIKE %s)")
        values.extend([f"%{mat}%", f"%{mat}%"])
    
    if conditions:
        where_clause = " OR ".join(conditions)
        sql = f"SELECT * FROM products WHERE {where_clause} LIMIT 12"
    else:
        sql = "SELECT * FROM products ORDER BY RANDOM() LIMIT 10"
        values = []
    
    try:
        cur.execute(sql, values)
        results = cur.fetchall()
        conn.close()
        
        if not results:
            return {
                "response": "Kh√¥ng t√¨m th·∫•y s·∫£n ph·∫©m ph√π h·ª£p.",
                "products": []
            }
        
        print(f"‚úÖ TIER 3 Success: Found {len(results)} products")
        return {
            "products": [dict(r) for r in results],
            "search_method": "keyword"
        }
    except Exception as e:
        conn.close()
        print(f"‚ùå TIER 3 failed: {e}")
        return {
            "response": "L·ªói t√¨m ki·∫øm.",
            "products": []
        }

def get_product_materials(headcode: str):
    """L·∫•y danh s√°ch v·∫≠t li·ªáu c·ªßa S·∫¢N PH·∫®M"""
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    cur.execute("SELECT product_name FROM products WHERE headcode = %s", (headcode,))
    prod = cur.fetchone()
    
    if not prod:
        conn.close()
        return {"response": f"‚ùå Kh√¥ng t√¨m th·∫•y s·∫£n ph·∫©m v·ªõi m√£ **{headcode}**"}
    
    sql = """
        SELECT 
            m.id_sap,
            m.material_name, 
            m.material_group,
            m.material_subgroup,
            m.material_subprice,
            m.unit as material_unit,
            m.image_url,
            pm.quantity, 
            pm.unit as pm_unit
        FROM product_materials pm
        INNER JOIN materials m ON pm.material_id_sap = m.id_sap
        WHERE pm.product_headcode = %s
        ORDER BY m.material_name ASC
    """
    
    try:
        cur.execute(sql, (headcode,))
        materials = cur.fetchall()
        print(f"üìä Found {len(materials)} materials for {headcode}")
    except Exception as e:
        print(f"‚ùå Query error: {e}")
        conn.close()
        return {"response": f"L·ªói truy v·∫•n database: {str(e)}"}
    
    conn.close()
    
    if not materials:
        return {
            "response": f"‚ö†Ô∏è S·∫£n ph·∫©m **{prod['product_name']}** ({headcode}) ch∆∞a c√≥ ƒë·ªãnh m·ª©c v·∫≠t li·ªáu.\n\n"
                        f" C√≥ th·ªÉ:\n"
                        f"‚Ä¢ S·∫£n ph·∫©m m·ªõi ch∆∞a nh·∫≠p ƒë·ªãnh m·ª©c\n"
                        f"‚Ä¢ Ch∆∞a import file product_materials.csv\n"
                        f"‚Ä¢ M√£ s·∫£n ph·∫©m trong product_materials kh√¥ng kh·ªõp\n\n"
                        f"Vui l√≤ng ki·ªÉm tra l·∫°i ho·∫∑c li√™n h·ªá b·ªô ph·∫≠n k·ªπ thu·∫≠t."
        }
    
    total = 0
    materials_with_price = []
    
    for mat in materials:
        latest_price = get_latest_material_price(mat['material_subprice'])
        quantity = float(mat['quantity']) if mat['quantity'] else 0.0 
        total_cost = quantity * latest_price
        total += total_cost
        
        materials_with_price.append({
            'id_sap': mat['id_sap'],
            'material_name': mat['material_name'],
            'material_group': mat['material_group'],
            'material_subgroup': mat['material_subgroup'],
            'material_unit': mat['material_unit'],
            'image_url': mat['image_url'],
            'quantity': quantity,
            'pm_unit': mat['pm_unit'],
            'price': latest_price,
            'unit_price': latest_price,
            'unit': mat['material_unit'],
            'total_cost': total_cost,
            'price_history': mat['material_subprice']
        })
    
    response = f"üéâ **ƒê·ªäNH M·ª®C V·∫¨T LI·ªÜU: {prod['product_name']}**\n"
    response += f"üè∑Ô∏è M√£: `{headcode}`\n"
    response += f"üì¶ T·ªïng s·ªë lo·∫°i v·∫≠t li·ªáu: **{len(materials_with_price)}**\n\n"
    response += "---\n\n"
    
    for idx, mat in enumerate(materials_with_price[:10], 1):
        response += f"**{idx}. {mat['material_name']}**\n"
        response += f"   ‚Ä¢ M√£ SAP: `{mat['id_sap']}`\n"
        response += f"   ‚Ä¢ Nh√≥m: {mat['material_group']}"
        if mat['material_subgroup']:
            response += f" - {mat['material_subgroup']}"
        response += f"\n"
        response += f"   ‚Ä¢ S·ªë l∆∞·ª£ng: {mat['quantity']} {mat['pm_unit']}\n"
        response += f"   ‚Ä¢ ƒê∆°n gi√° m·ªõi nh·∫•t: {mat['unit_price']:,.2f} VNƒê\n"
        response += f"   ‚Ä¢ Th√†nh ti·ªÅn: **{mat['total_cost']:,.2f} VNƒê**\n"
        
        if mat.get('image_url'):
            response += f"   ‚Ä¢ [üì∑ Xem ·∫£nh]({mat['image_url']})\n"
        
        response += "\n"
    
    if len(materials_with_price) > 10:
        response += f"\n*...v√† {len(materials_with_price)-10} v·∫≠t li·ªáu kh√°c.*\n"
    
    response += f"\n---\n\nüí∞ **T·ªîNG CHI PH√ç NGUY√äN V·∫¨T LI·ªÜU: {total:,.2f} VNƒê**"
    response += f"\n\n‚ö†Ô∏è **L∆∞u √Ω:** Gi√° ƒë∆∞·ª£c t√≠nh t·ª´ l·ªãch s·ª≠ mua h√†ng g·∫ßn nh·∫•t. Gi√° th·ª±c t·∫ø c√≥ th·ªÉ thay ƒë·ªïi."
    
    return {
        "response": response,
        "materials": materials_with_price,
        "total_cost": total,
        "product_name": prod['product_name']
    }

# ========================================
# FIX BUG TRONG chatbot_api.py
# D√≤ng 1230 - 1260 (h√†m calculate_product_cost)
# ========================================
# ========================================
# FIX 3: D√íNG ~1288-1370
# Thay th·∫ø h√†m calculate_product_cost
# ========================================

def calculate_product_cost(headcode: str):
    """T√≠nh CHI PH√ç NGUY√äN V·∫¨T LI·ªÜU s·∫£n ph·∫©m (ƒê∆°n gi·∫£n h√≥a V4.7)"""
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    cur.execute("SELECT product_name, category FROM products WHERE headcode = %s", (headcode,))
    prod = cur.fetchone()
    
    if not prod:
        conn.close()
        return {"response": f"‚ùå Kh√¥ng t√¨m th·∫•y s·∫£n ph·∫©m v·ªõi m√£ **{headcode}**"}
    
    sql = """
        SELECT 
            m.material_name,
            m.material_group,
            m.material_subprice,
            m.unit as material_unit,
            pm.quantity,
            pm.unit as pm_unit
        FROM product_materials pm
        INNER JOIN materials m ON pm.material_id_sap = m.id_sap
        WHERE pm.product_headcode = %s
        ORDER BY m.material_name ASC
    """
    
    try:
        cur.execute(sql, (headcode,))
        materials = cur.fetchall()
        print(f"üí∞ Cost calculation for {headcode}: {len(materials)} materials")
    except Exception as e:
        print(f"‚ùå Query error: {e}")
        conn.close()
        return {"response": f"L·ªói truy v·∫•n database: {str(e)}"}
    
    conn.close()
    
    if not materials:
        return {
            "response": f"‚ö†Ô∏è S·∫£n ph·∫©m **{prod['product_name']}** ({headcode}) ch∆∞a c√≥ ƒë·ªãnh m·ª©c v·∫≠t li·ªáu.\n\n"
                       f"**Nguy√™n nh√¢n c√≥ th·ªÉ:**\n"
                       f"‚Ä¢ S·∫£n ph·∫©m m·ªõi ch∆∞a nh·∫≠p ƒë·ªãnh m·ª©c\n"
                       f"‚Ä¢ Ch∆∞a import file `product_materials.csv`\n"
                       f"‚Ä¢ M√£ s·∫£n ph·∫©m trong file CSV kh√¥ng kh·ªõp v·ªõi `{headcode}`\n\n"
                       f"**Gi·∫£i ph√°p:**\n"
                       f"1. Ki·ªÉm tra file CSV c√≥ d√≤ng n√†o v·ªõi `product_headcode = {headcode}`\n"
                       f"2. Import l·∫°i file qua sidebar: **Import D·ªØ Li·ªáu ‚Üí ƒê·ªãnh M·ª©c**"
        }
    
    # ‚úÖ T√≠nh T·ªîNG CHI PH√ç V·∫¨T LI·ªÜU
    material_cost = 0.0
    material_count = len(materials)
    materials_detail = []
    
    for mat in materials:
        quantity = float(mat['quantity']) if mat['quantity'] else 0.0
        latest_price = get_latest_material_price(mat['material_subprice'])
        total_cost = quantity * latest_price
        material_cost += total_cost
        
        materials_detail.append({
            'name': mat['material_name'],
            'group': mat['material_group'],
            'quantity': quantity,
            'unit': mat['pm_unit'],
            'unit_price': latest_price,
            'total': total_cost
        })
    
    # ‚úÖ RESPONSE ƒê∆†N GI·∫¢N - CH·ªà CHI PH√ç V·∫¨T LI·ªÜU
    response = f"""
üí∞ **B√ÅO GI√Å NGUY√äN V·∫¨T LI·ªÜU**

üì¶ **S·∫£n ph·∫©m:** {prod['product_name']}
üè∑Ô∏è **M√£:** `{headcode}`
üìÇ **Danh m·ª•c:** {prod['category'] or 'N/A'}

---

**CHI TI·∫æT NGUY√äN V·∫¨T LI·ªÜU ({material_count} lo·∫°i):**

"""
    
    for idx, mat in enumerate(materials_detail[:15], 1):
        response += f"{idx}. **{mat['name']}** ({mat['group']})\n"
        response += f"   ‚Ä¢ S·ªë l∆∞·ª£ng: {mat['quantity']} {mat['unit']}\n"
        response += f"   ‚Ä¢ ƒê∆°n gi√°: {mat['unit_price']:,.0f} VNƒê\n"
        response += f"   ‚Ä¢ Th√†nh ti·ªÅn: **{mat['total']:,.0f} VNƒê**\n\n"
    
    if len(materials_detail) > 15:
        response += f"*...v√† {len(materials_detail)-15} v·∫≠t li·ªáu kh√°c*\n\n"
    
    response += f"---\n\n"
    response += f"‚úÖ **T·ªîNG CHI PH√ç NGUY√äN V·∫¨T LI·ªÜU: {material_cost:,.0f} VNƒê**\n\n"
    response += f"üìã **L∆∞u √Ω:** Gi√° ƒë∆∞·ª£c t√≠nh t·ª´ l·ªãch s·ª≠ mua h√†ng g·∫ßn nh·∫•t.\n"
    response += f"üí° **Mu·ªën xem chi ti·∫øt ƒë·ªãnh m·ª©c?** H·ªèi: _\"Ph√¢n t√≠ch v·∫≠t li·ªáu {headcode}\"_"
    
    return {
        "response": response,
        "material_cost": material_cost,
        "material_count": material_count,
        "materials": materials_detail,
        "suggested_prompts":[
            "Ph√¢n t√≠ch v·∫≠t li·ªáu {headcode}"
        ]
    }

# ========================================
# MATERIAL FUNCTIONS
# ========================================

def search_materials(params: Dict):
    """T√¨m ki·∫øm NGUY√äN V·∫¨T LI·ªÜU v·ªõi gi√° t·ª´ material_subprice"""
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    query_parts = []
    if params.get("material_name"): 
        query_parts.append(params["material_name"])
    if params.get("material_group"): 
        query_parts.append(params["material_group"])
    if params.get("usage_context"): 
        query_parts.append(params["usage_context"])
    if params.get("keywords_vector"): 
        query_parts.append(params["keywords_vector"])
    
    query_text = " ".join(query_parts) if query_parts else "v·∫≠t li·ªáu n·ªôi th·∫•t"
    print(f"üîç Searching materials for: {query_text}")
    
    query_vector = generate_embedding(query_text)
    
    if query_vector:
        try:
            filter_clause = "1=1"
            filter_params = []
            
            if params.get("material_group"):
                filter_clause = "material_group ILIKE %s"
                filter_params = [f"%{params['material_group']}%"]
            
            sql = f"""
                SELECT 
                    id_sap, material_name, material_group, material_subgroup,
                    material_subprice, unit, image_url,
                    (description_embedding <=> %s::vector) as distance
                FROM materials
                WHERE description_embedding IS NOT NULL AND {filter_clause}
                ORDER BY distance ASC
                LIMIT 10
            """
            
            cur.execute(sql, [query_vector] + filter_params)
            results = cur.fetchall()
            
            if results:
                print(f"‚úÖ Vector search: Found {len(results)} materials")
                
                materials_with_price = []
                for mat in results:
                    mat_dict = dict(mat)
                    mat_dict['price'] = get_latest_material_price(mat_dict['material_subprice'])
                    materials_with_price.append(mat_dict)
                
                conn.close()
                return {
                    "materials": materials_with_price,
                    "search_method": "vector"
                }
        except Exception as e:
            print(f"‚ö†Ô∏è Vector search failed: {e}")
    
    print("‚ÑπÔ∏è Keyword search for materials")
    conditions = []
    values = []
    
    if params.get("material_name"):
        name = params['material_name']
        conditions.append("(material_name ILIKE %s OR material_group ILIKE %s)")
        values.extend([f"%{name}%", f"%{name}%"])
    
    if params.get("material_group"):
        group = params['material_group']
        conditions.append("material_group ILIKE %s")
        values.append(f"%{group}%")
    
    if conditions:
        where_clause = " OR ".join(conditions)
        sql = f"SELECT * FROM materials WHERE {where_clause} LIMIT 15"
    else:
        sql = "SELECT * FROM materials ORDER BY material_name ASC LIMIT 10"
        values = []
    
    try:
        cur.execute(sql, values)
        results = cur.fetchall()
        conn.close()
        
        if not results:
            return {
                "response": "Kh√¥ng t√¨m th·∫•y v·∫≠t li·ªáu ph√π h·ª£p.",
                "materials": []
            }
        
        materials_with_price = []
        for mat in results:
            mat_dict = dict(mat)
            mat_dict['price'] = get_latest_material_price(mat.get('material_subprice'))
            materials_with_price.append(mat_dict)
        
        print(f"‚úÖ Keyword search: Found {len(materials_with_price)} materials")
        return {
            "materials": materials_with_price,
            "search_method": "keyword"
        }
    except Exception as e:
        conn.close()
        print(f"‚ùå Material search failed: {e}")
        return {
            "response": "L·ªói t√¨m ki·∫øm v·∫≠t li·ªáu.",
            "materials": []
        }

def get_material_detail(id_sap: str = None, material_name: str = None):
    """Xem chi ti·∫øt V·∫¨T LI·ªÜU + l·ªãch s·ª≠ gi√° + s·∫£n ph·∫©m s·ª≠ d·ª•ng"""
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    if id_sap:
        cur.execute("SELECT * FROM materials WHERE id_sap = %s", (id_sap,))
    elif material_name:
        cur.execute("SELECT * FROM materials WHERE material_name ILIKE %s LIMIT 1", (f"%{material_name}%",))
    else:
        conn.close()
        return {"response": "‚ö†Ô∏è C·∫ßn cung c·∫•p m√£ SAP ho·∫∑c t√™n v·∫≠t li·ªáu."}
    
    material = cur.fetchone()
    
    if not material:
        conn.close()
        return {"response": f"‚ùå Kh√¥ng t√¨m th·∫•y v·∫≠t li·ªáu **{id_sap or material_name}**"}
    
    latest_price = get_latest_material_price(material['material_subprice'])

    sql = """
        SELECT 
            p.headcode,
            p.product_name,
            p.category,
            p.sub_category,
            p.project,
            pm.quantity,
            pm.unit
        FROM product_materials pm
        INNER JOIN products p ON pm.product_headcode = p.headcode
        WHERE pm.material_id_sap = %s
        ORDER BY p.product_name ASC
        LIMIT 20
    """
    
    try:
        cur.execute(sql, (material['id_sap'],))
        used_in_products = cur.fetchall()
        print(f"üîó Material {material['id_sap']} used in {len(used_in_products)} products")
    except Exception as e:
        print(f"‚ùå Query error: {e}")
        used_in_products = []
    
    try:
        cur.execute("""
            SELECT 
                COUNT(DISTINCT pm.product_headcode) as product_count,
                COUNT(DISTINCT p.project) as project_count,
                SUM(pm.quantity) as total_quantity
            FROM product_materials pm
            LEFT JOIN products p ON pm.product_headcode = p.headcode
            WHERE pm.material_id_sap = %s
        """, (material['id_sap'],))
        stats = cur.fetchone()
    except Exception as e:
        print(f"‚ùå Stats query error: {e}")
        stats = {
            'product_count': 0,
            'project_count': 0,
            'total_quantity': 0
        }
    
    conn.close()
    
    price_history = []
    try:
        if material['material_subprice']:
            price_history = json.loads(material['material_subprice'])
    except:
        pass
    
    response = f"üß± **CHI TI·∫æT NGUY√äN V·∫¨T LI·ªÜU**\n\n"
    response += f"üì¶ **T√™n:** {material['material_name']}\n\n"
    response += f"üè∑Ô∏è **M√£ SAP:** `{material['id_sap']}\n\n"
    response += f"üìÇ **Nh√≥m:** {material['material_group']}\n\n"
    
    if material.get('material_subgroup'):
        response += f" - {material['material_subgroup']}"
    
    response += f"üí∞ **Gi√° m·ªõi nh·∫•t:** {latest_price:,.2f} VNƒê/{material['unit']}\n\n"
    response += f"üìä **TH·ªêNG K√ä S·ª¨ D·ª§NG:**\n"
    response += f"‚Ä¢ ƒê∆∞·ª£c s·ª≠ d·ª•ng trong **{stats['product_count']} s·∫£n ph·∫©m**\n"
    response += f"‚Ä¢ Xu·∫•t hi·ªán ·ªü **{stats['project_count']} d·ª± √°n**\n"
    response += f"‚Ä¢ T·ªïng s·ªë l∆∞·ª£ng: **{stats.get('total_quantity', 0) or 0} {material['unit']}**\n\n"
    response += "\n---\n\n"
    
    if price_history and len(price_history) > 0:
        response += "üìà **L·ªäCH S·ª¨ GI√Å:**\n\n"
        for idx, ph in enumerate(sorted(price_history, key=lambda x: x['date'], reverse=True)[:5], 1):
            response += f"{idx}. **{ph['date']}**: {ph['price']:,.2f} VNƒê\n"
        response += "\n---\n\n"
    
    if used_in_products and len(used_in_products) > 0:
        response += f"üîó **C√ÅC S·∫¢N PH·∫®M S·ª¨ D·ª§NG V·∫¨T LI·ªÜU N√ÄY:**\n\n"
        
        for idx, prod in enumerate(used_in_products[:10], 1):
            response += f"{idx}. **{prod['product_name']}** (`{prod['headcode']}`)\n"
            response += f"   ‚Ä¢ Danh m·ª•c: {prod.get('category', 'N/A')}"
            if prod.get('sub_category'):
                response += f" - {prod['sub_category']}"
            response += "\n"
            
            if prod.get('project'):
                response += f"   ‚Ä¢ D·ª± √°n: {prod['project']}\n"
            
            response += f"   ‚Ä¢ S·ª≠ d·ª•ng: **{prod['quantity']} {prod['unit']}**\n\n"
        
        if len(used_in_products) > 10:
            response += f"*...v√† {len(used_in_products)-10} s·∫£n ph·∫©m kh√°c*\n\n"
    else:
        response += "üîó **CH∆ØA C√ì S·∫¢N PH·∫®M S·ª¨ D·ª§NG**\n\n"
        response += "_V·∫≠t li·ªáu n√†y ch∆∞a ƒë∆∞·ª£c g·∫Øn v√†o s·∫£n ph·∫©m n√†o trong h·ªá th·ªëng._\n\n"
    
    if material.get('image_url'):
        response += f"---\n\nüñºÔ∏è **Xem ·∫£nh v·∫≠t li·ªáu:** [Google Drive Link]({material['image_url']})\n"
        response += f"_(Click ƒë·ªÉ xem ·∫£nh chi ti·∫øt)_"
    
    return {
        "response": response,
        # "material_detail": dict(material),
        "materials": [{  # ‚úÖ ƒê·ªïi th√†nh list gi·ªëng search_materials
        **dict(material),
        'price': latest_price  # ‚úÖ Th√™m key 'price'
    }],
        "latest_price": latest_price,
        "price_history": price_history,
        "used_in_products": [dict(p) for p in used_in_products],
        "stats": dict(stats) if stats else {},
        "has_image": bool(material.get('image_url'))
    }



def list_material_groups():
    """Li·ªát k√™ c√°c nh√≥m v·∫≠t li·ªáu v·ªõi gi√° t√≠nh t·ª´ material_subprice"""
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    sql = """
        SELECT 
            material_group,
            COUNT(*) as count,
            array_agg(DISTINCT material_subprice) as all_prices
        FROM materials
        WHERE material_group IS NOT NULL
        GROUP BY material_group
        ORDER BY count DESC
    """
    cur.execute(sql)
    groups = cur.fetchall()
    conn.close()
    
    if not groups:
        return {"response": "Ch∆∞a c√≥ d·ªØ li·ªáu nh√≥m v·∫≠t li·ªáu."}
    
    response = f"üìã **DANH S√ÅCH NH√ìM V·∫¨T LI·ªÜU ({len(groups)} nh√≥m):**\n\n"
    
    groups_with_stats = []
    for g in groups:
        prices = []
        for price_json in g['all_prices']:
            if price_json:
                latest = get_latest_material_price(price_json)
                if latest > 0:
                    prices.append(latest)
        
        avg_price = sum(prices) / len(prices) if prices else 0
        min_price = min(prices) if prices else 0
        max_price = max(prices) if prices else 0
        
        groups_with_stats.append({
            'material_group': g['material_group'],
            'count': g['count'],
            'avg_price': avg_price,
            'min_price': min_price,
            'max_price': max_price
        })
    
    for idx, g in enumerate(groups_with_stats, 1):
        response += f"{idx}. **{g['material_group']}** ({g['count']} lo·∫°i)\n"
        if g['avg_price'] > 0:
            response += f"   ‚Ä¢ Gi√° TB: {g['avg_price']:,.2f} VNƒê\n"
            response += f"   ‚Ä¢ Kho·∫£ng gi√°: {g['min_price']:,.2f} - {g['max_price']:,.2f} VNƒê\n"
        response += "\n"
    
    return {
        "response": response,
        "material_groups": groups_with_stats
    }

# ========================================
# MAIN CHAT ENDPOINT
# ========================================

@app.post("/chat")
def chat(msg: ChatMessage):
    """Main chat logic"""
    try:
        user_message = msg.message
        context = msg.context or {}
        
        intent_data = get_intent_and_params(user_message, context)
        
        if intent_data.get("intent") == "error":
            return {"response": "Xin l·ªói, h·ªá th·ªëng ƒëang b·∫≠n. Vui l√≤ng th·ª≠ l·∫°i."}
        
        intent = intent_data["intent"]
        params = intent_data.get("params", {})
        
        result_response = None
        result_count = 0
        
        # GREETING
        if intent == "greeting":
            result_response = {
                "response": "üëã Xin ch√†o! T√¥i l√† tr·ª£ l√Ω AI c·ªßa AA Corporation.\n\n"
                           "T√¥i c√≥ th·ªÉ gi√∫p b·∫°n:\n"
                           "‚Ä¢ üîç **T√¨m s·∫£n ph·∫©m** (b√†n, gh·∫ø, sofa...)\n"
                           "‚Ä¢ üß± **T√¨m nguy√™n v·∫≠t li·ªáu** (g·ªó, da, ƒë√°, v·∫£i...)\n"
                           "‚Ä¢ üí∞ **T√≠nh chi ph√≠** s·∫£n ph·∫©m\n"
                           "‚Ä¢ üìã **Xem ƒë·ªãnh m·ª©c** nguy√™n v·∫≠t li·ªáu\n\n"
                           "B·∫°n c·∫ßn t√¨m g√¨ h√¥m nay?",
                
                "suggested_prompts": [
                    "üîç T√¨m s·∫£n ph·∫©m", 
                    "üß± T√¨m nguy√™n v·∫≠t li·ªáu", 
                    "üí∞ Xem gi√° s·∫£n ph·∫©m",
                    "üìã Danh s√°ch nh√≥m v·∫≠t li·ªáu"
                ]
            }
        
        # PRODUCT FLOW
        # elif intent == "search_product":
        #     search_result = search_products(params)
        #     products = search_result.get("products", [])
        #     result_count = len(products)
            
        #     if not products:
        #         result_response = {"response": search_result.get("response", "Kh√¥ng t√¨m th·∫•y s·∫£n ph·∫©m.")}
        #     else:
        #         response_text = ""
        #         suggested_prompts = []
                
        #         if intent_data.get("is_broad_query"):
        #             follow_up = intent_data.get("follow_up_question", "B·∫°n mu·ªën t√¨m lo·∫°i c·ª• th·ªÉ n√†o?")
        #             response_text = (
        #                 f"üîé T√¨m th·∫•y **{len(products)} s·∫£n ph·∫©m** ph√π h·ª£p v·ªõi t·ª´ kh√≥a chung.\n"
        #                 f"*(T√¥i ƒë√£ ch·ªçn l·ªçc c√°c m·∫´u ph·ªï bi·∫øn nh·∫•t b√™n d∆∞·ªõi)*\n\n"
        #                 f"üí° **G·ª£i √Ω:** {follow_up}"
        #             )
        #             actions = intent_data.get("suggested_actions", [])
        #             suggested_prompts = [f"üîç {a}" for a in actions] if actions else []
        #         else:
        #             response_text = f"‚úÖ ƒê√£ t√¨m th·∫•y **{len(products)} s·∫£n ph·∫©m** ƒë√∫ng y√™u c·∫ßu c·ªßa b·∫°n."
        #             suggested_prompts = [
        #                 f"üí∞ T√≠nh chi ph√≠ {products[0]['headcode']}",
        #                 f"üìã Xem v·∫≠t li·ªáu {products[0]['headcode']}"
        #             ]
                
        #         result_response = {
        #             "response": response_text,
        #             "products": products,
        #             "suggested_prompts": suggested_prompts
        #         }
        #         # CROSS-TABLE: T√¨m s·∫£n ph·∫©m theo v·∫≠t li·ªáu
        
        # PRODUCT FLOW - C·∫¨P NH·∫¨T V4.8 (Feedback Ranking)




        elif intent == "search_product":
            search_result = search_products(params, session_id=msg.session_id)
            products = search_result.get("products", [])
            
            # ‚úÖ search_products ƒë√£ x·ª≠ l√Ω H·∫æT ranking r·ªìi, kh√¥ng c·∫ßn g·ªçi g√¨ th√™m
            
            ranking_summary = search_result.get("ranking_summary", {})
            result_count = len(products)
            
            if not products:
                result_response = {"response": search_result.get("response", "Kh√¥ng t√¨m th·∫•y s·∫£n ph·∫©m.")}
            else:





                response_text = ""
                suggested_prompts = []
                
                if intent_data.get("is_broad_query"):
                    follow_up = intent_data.get("follow_up_question", "B·∫°n mu·ªën t√¨m lo·∫°i c·ª• th·ªÉ n√†o?")
                    response_text = (
                        f"üîé T√¨m th·∫•y **{len(products)} s·∫£n ph·∫©m** ph√π h·ª£p v·ªõi t·ª´ kh√≥a chung.\n"
                        f"*(T√¥i ƒë√£ ch·ªçn l·ªçc c√°c m·∫´u ph·ªï bi·∫øn nh·∫•t b√™n d∆∞·ªõi)*\n\n"
                        f"üí° **G·ª£i √Ω:** {follow_up}"
                    )
                    actions = intent_data.get("suggested_actions", [])
                    suggested_prompts = [f"üîç {a}" for a in actions] if actions else []
                else:
                    response_text = f"‚úÖ ƒê√£ t√¨m th·∫•y **{len(products)} s·∫£n ph·∫©m** ƒë√∫ng y√™u c·∫ßu c·ªßa b·∫°n."
                    
                    # ‚úÖ TH√äM: Hi·ªÉn th·ªã th√¥ng tin ranking n·∫øu c√≥
                    if ranking_summary['ranking_applied']:
                        response_text += f"\n\n‚≠ê **{ranking_summary['boosted_items']} s·∫£n ph·∫©m** ƒë∆∞·ª£c ∆∞u ti√™n d·ª±a tr√™n l·ªãch s·ª≠ t√¨m ki·∫øm."
                    
                    suggested_prompts = [
                        f"üí∞ T√≠nh chi ph√≠ {products[0]['headcode']}",
                        f"üìã Xem v·∫≠t li·ªáu {products[0]['headcode']}"
                    ]
                
                result_response = {
                    "response": response_text,
                    "products": products,
                    "suggested_prompts": suggested_prompts,
                    "ranking_summary": ranking_summary,  # ‚úÖ TH√äM
                    "can_provide_feedback": True  # ‚úÖ TH√äM
                }
        
        
        elif intent == "search_product_by_material":
            material_query = params.get("material_name") or params.get("material_primary") or params.get("keywords_vector")
            
            if not material_query:
                result_response = {
                    "response": "‚ö†Ô∏è B·∫°n mu·ªën t√¨m s·∫£n ph·∫©m l√†m t·ª´ v·∫≠t li·ªáu n√†o?",
                    "suggested_prompts": [
                        "üîç B√†n l√†m t·ª´ ƒë√° marble",
                        "üîç Gh·∫ø g·ªó teak",
                        "üîç T·ªß g·ªó s·ªìi"
                    ]
                }
            else:
                search_result = search_products_by_material(material_query, params)
                products = search_result.get("products", [])
                
                feedback_scores = get_feedback_boost_for_query(user_message, "product")
                if feedback_scores:
                    products = rerank_with_feedback(products, feedback_scores, "headcode")
                
                result_count = len(products)
                
                if not products:
                    matched_mats = search_result.get("matched_materials", [])
                    result_response = {
                        "response": f"üîç ƒê√£ t√¨m th·∫•y v·∫≠t li·ªáu: **{', '.join(matched_mats)}**\n\n"
                                   f"Nh∆∞ng kh√¥ng c√≥ s·∫£n ph·∫©m n√†o s·ª≠ d·ª•ng v·∫≠t li·ªáu n√†y trong h·ªá th·ªëng.\n\n"
                                   f"üí° Th·ª≠ t√¨m ki·∫øm kh√°c ho·∫∑c m·ªü r·ªông ƒëi·ªÅu ki·ªán.",
                        "materials": []
                    }
                else:
                    explanation = search_result.get("explanation", "")
                    response_text = f"‚úÖ {explanation}\n\n"
                    response_text += f"üì¶ T√¨m th·∫•y **{len(products)} s·∫£n ph·∫©m**:"
                    
                    result_response = {
                        "response": response_text,
                        "products": products,
                        "search_method": "cross_table",
                        "can_provide_feedback": True
                    }

        
    


        # CROSS-TABLE: T√¨m v·∫≠t li·ªáu cho s·∫£n ph·∫©m
        elif intent == "search_material_for_product":
            # 1. L·∫•y query t·ª´ params ho·∫∑c context
            product_query = params.get("category") or params.get("usage_context") or params.get("keywords_vector")
            
            if not product_query:
                result_response = {
                    "response": "‚ö†Ô∏è B·∫°n mu·ªën t√¨m v·∫≠t li·ªáu ƒë·ªÉ l√†m s·∫£n ph·∫©m g√¨?",
                    "suggested_prompts": [
                        "üß± V·∫≠t li·ªáu l√†m b√†n ƒÉn",
                        "üß± Nguy√™n li·ªáu gh·∫ø sofa",
                        "üß± ƒê√° l√†m b√†n coffee"
                    ]
                }
            else:
                # 2. G·ªçi h√†m t√¨m ki·∫øm
                search_result = search_materials_for_product(product_query, params)
                materials = search_result.get("materials", [])
                
                # 3. [M·ªöI] √Åp d·ª•ng Feedback Ranking (Gi·ªëng Intent 3)
                # D√πng query g·ªëc c·ªßa user ƒë·ªÉ t√¨m feedback t∆∞∆°ng t·ª±
                feedback_scores = get_feedback_boost_for_query(user_message, "material")
                if feedback_scores:
                    materials = rerank_with_feedback(materials, feedback_scores, "id_sap")
                
                # 4. [M·ªöI] L·∫•y th√¥ng tin Ranking Summary ƒë·ªÉ hi·ªÉn th·ªã UI
                ranking_summary = get_ranking_summary(materials)
                
                result_count = len(materials)
                
                if not materials:
                    result_response = {
                        "response": "Kh√¥ng t√¨m th·∫•y v·∫≠t li·ªáu ph√π h·ª£p.",
                        "materials": []
                    }
                else:
                    explanation = search_result.get("explanation", "")
                    
                    response_text = f"‚úÖ {explanation}\n\n"
                    
                    # Hi·ªÉn th·ªã th√¥ng b√°o n·∫øu c√≥ Ranking
                    if ranking_summary['ranking_applied']:
                         response_text += f"‚≠ê **{ranking_summary['boosted_items']} v·∫≠t li·ªáu** ƒë∆∞·ª£c ∆∞u ti√™n d·ª±a tr√™n l·ªãch s·ª≠.\n\n"
                         
                    response_text += f"üß± T√¨m th·∫•y **{len(materials)} v·∫≠t li·ªáu** th∆∞·ªùng d√πng:\n\n"
                    
                    for idx, mat in enumerate(materials[:5], 1):
                        response_text += f"{idx}. **{mat['material_name']}**\n"
                        response_text += f"   ‚Ä¢ Nh√≥m: {mat['material_group']}\n"
                        response_text += f"   ‚Ä¢ Gi√°: {mat.get('price', 0):,.0f} VNƒê/{mat.get('unit', '')}\n"
                        response_text += f"   ‚Ä¢ D√πng trong {mat.get('usage_count', 0)} s·∫£n ph·∫©m\n\n"
                    
                    result_response = {
                        "response": response_text,
                        "materials": materials,
                        "search_method": "cross_table_product_to_material", # ƒê√°nh d·∫•u ƒë·ªÉ UI nh·∫≠n bi·∫øt
                        "ranking_summary": ranking_summary,   # Truy·ªÅn xu·ªëng UI
                        "can_provide_feedback": True          # B·∫≠t n√∫t Feedback
                    }

        



        elif intent == "query_product_materials":
            headcode = params.get("headcode")
            
            if not headcode and context.get("last_search_results"):
                headcode = context["last_search_results"][0]
                
            if not headcode:
                result_response = {
                    "response": "‚ö†Ô∏è B·∫°n mu·ªën xem v·∫≠t li·ªáu c·ªßa s·∫£n ph·∫©m n√†o? Vui l√≤ng cung c·∫•p m√£ ho·∫∑c t√¨m ki·∫øm s·∫£n ph·∫©m tr∆∞·ªõc.",
                    "suggested_prompts": ["üîç T√¨m gh·∫ø sofa", "üîç T√¨m b√†n ƒÉn"]
                }
            else:
                result_response = get_product_materials(headcode)
                result_count = len(result_response.get("materials", []))
        
        elif intent == "calculate_product_cost":
            headcode = params.get("headcode")
            
            if not headcode and context.get("last_search_results"):
                headcode = context["last_search_results"][0]
            
            if not headcode:
                result_response = {
                    "response": "‚ö†Ô∏è B·∫°n mu·ªën xem chi ph√≠ s·∫£n ph·∫©m n√†o? Vui l√≤ng cung c·∫•p m√£ ho·∫∑c t√¨m ki·∫øm s·∫£n ph·∫©m tr∆∞·ªõc.",
                    "suggested_prompts": ["üîç T√¨m gh·∫ø sofa", "üîç T√¨m b√†n ƒÉn"]
                }
            else:
                result_response = calculate_product_cost(headcode)
        
        # MATERIAL FLOW
        # elif intent == "search_material":
        #     search_result = search_materials(params)
        #     materials = search_result.get("materials", [])
        #     result_count = len(materials)
            
        #     if not materials:
        #         result_response = {
        #             "response": search_result.get("response", "Kh√¥ng t√¨m th·∫•y v·∫≠t li·ªáu ph√π h·ª£p."),
        #             "materials": []
        #         }
        #     else:
        #         response_text = ""
                
        #         if intent_data.get("is_broad_query"):
        #             follow_up = intent_data.get("follow_up_question", "B·∫°n c·∫ßn t√¨m lo·∫°i v·∫≠t li·ªáu c·ª• th·ªÉ n√†o?")
        #             response_text = (
        #                 f"üîé T√¨m th·∫•y **{len(materials)} nguy√™n v·∫≠t li·ªáu** ph√π h·ª£p.\n\n"
        #                 f"üí° **G·ª£i √Ω:** {follow_up}"
        #             )
        #         else:
        #             response_text = f"‚úÖ ƒê√£ t√¨m th·∫•y **{len(materials)} nguy√™n v·∫≠t li·ªáu** ƒë√∫ng y√™u c·∫ßu."
                
        #         response_text += "\n\nüì¶ **K·∫æT QU·∫¢:**\n"
        #         for idx, mat in enumerate(materials[:8], 1):
        #             response_text += f"\n{idx}. **{mat['material_name']}**"
        #             response_text += f"\n   ‚Ä¢ M√£: `{mat['id_sap']}`"
        #             response_text += f"\n   ‚Ä¢ Nh√≥m: {mat['material_group']}"
        #             response_text += f"\n   ‚Ä¢ Gi√°: {mat.get('price', 0):,.2f} VNƒê/{mat.get('unit', '')}"
        #             if mat.get('image_url'):
        #                 response_text += f"\n   ‚Ä¢ [üì∑ Xem ·∫£nh]({mat['image_url']})"
                
        #         if len(materials) > 8:
        #             response_text += f"\n\n*...v√† {len(materials)-8} v·∫≠t li·ªáu kh√°c*"
                
        #         suggested_prompts = []
        #         if materials:
        #             first_mat = materials[0]
        #             suggested_prompts = [
        #                 f"üîç Chi ti·∫øt {first_mat['material_name']}",
        #                 "üìã Xem nh√≥m v·∫≠t li·ªáu kh√°c"
        #             ]
                
        #         result_response = {
        #             "response": response_text,
        #             "materials": materials,
        #             "suggested_prompts": suggested_prompts
        #         }
        
      
# MATERIAL FLOW - C·∫¨P NH·∫¨T V4.8 (Feedback Ranking)
        elif intent == "search_material":
            search_result = search_materials(params)
            materials = search_result.get("materials", [])
            
            # üÜï √ÅP D·ª§NG FEEDBACK RANKING
            materials = apply_feedback_to_search(
                materials,
                user_message,
                search_type="material",
                id_key="id_sap"
            )
            
            # üÜï L·∫•y ranking summary
            ranking_summary = get_ranking_summary(materials)
            
            result_count = len(materials)
            
            if not materials:
                result_response = {
                    "response": search_result.get("response", "Kh√¥ng t√¨m th·∫•y v·∫≠t li·ªáu ph√π h·ª£p."),
                    "materials": []
                }
            else:
                response_text = ""
                
                if intent_data.get("is_broad_query"):
                    follow_up = intent_data.get("follow_up_question", "B·∫°n c·∫ßn t√¨m lo·∫°i v·∫≠t li·ªáu c·ª• th·ªÉ n√†o?")
                    response_text = (
                        f"üîé T√¨m th·∫•y **{len(materials)} nguy√™n v·∫≠t li·ªáu** ph√π h·ª£p.\n\n"
                        f"üíñ **Ghi ch√∫:** {follow_up}"
                    )
                else:
                    response_text = f"‚úÖ ƒê√£ t√¨m th·∫•y **{len(materials)} nguy√™n v·∫≠t li·ªáu** ƒë√∫ng y√™u c·∫ßu."
                    
                    # üÜï Hi·ªÉn th·ªã ranking info
                    if ranking_summary['ranking_applied']:
                        response_text += f"\n\n‚≠ê **{ranking_summary['boosted_items']} v·∫≠t li·ªáu** ƒë∆∞·ª£c ∆∞u ti√™n."
                
                response_text += "\n\nüì¶ **K·∫æT QU·∫¢:**\n"
                for idx, mat in enumerate(materials[:8], 1):
                    response_text += f"\n{idx}. **{mat['material_name']}**"
                    response_text += f"\n   ‚Ä¢ M√£: `{mat['id_sap']}`"
                    response_text += f"\n   ‚Ä¢ Nh√≥m: {mat['material_group']}"
                    response_text += f"\n   ‚Ä¢ Gi√°: {mat.get('price', 0):,.2f} VNƒê/{mat.get('unit', '')}"
                    
                    # üÜï Hi·ªÉn th·ªã feedback indicator
                    if mat.get('has_feedback'):
                        response_text += f"\n   ‚≠ê {mat['feedback_count']} ng∆∞·ªùi ƒë√£ ch·ªçn"
                    
                    if mat.get('image_url'):
                        response_text += f"\n   ‚Ä¢ [üì∑ Xem ·∫£nh]({mat['image_url']})"
                
                if len(materials) > 8:
                    response_text += f"\n\n*...v√† {len(materials)-8} v·∫≠t li·ªáu kh√°c*"
                
                suggested_prompts = []
                if materials:
                    first_mat = materials[0]
                    suggested_prompts = [
                        f"üîç Chi ti·∫øt {first_mat['material_name']}",
                        "üìã Xem nh√≥m v·∫≠t li·ªáu kh√°c"
                    ]
                
                result_response = {
                    "response": response_text,
                    "materials": materials,
                    "suggested_prompts": suggested_prompts,
                    "ranking_summary": ranking_summary,  # üÜï
                    "can_provide_feedback": True  # üÜï
                }      

        elif intent == "query_material_detail":
            id_sap = params.get("id_sap")
            material_name = params.get("material_name")
            
            if not id_sap and not material_name and context.get("current_materials"):
                first_mat = context["current_materials"][0]
                id_sap = first_mat.get("id_sap")
            
            if not id_sap and not material_name:
                result_response = {
                    "response": "‚ö†Ô∏è B·∫°n mu·ªën xem chi ti·∫øt v·∫≠t li·ªáu n√†o? Vui l√≤ng cung c·∫•p m√£ SAP ho·∫∑c t√™n v·∫≠t li·ªáu.",
                    "suggested_prompts": ["üß± T√¨m g·ªó s·ªìi", "üìã Danh s√°ch nh√≥m v·∫≠t li·ªáu"]
                }
            else:
                result_response = get_material_detail(id_sap=id_sap, material_name=material_name)
                result_count = len(result_response.get("used_in_products", []))
        
        elif intent == "list_material_groups":
            result_response = list_material_groups()
        
        # UNKNOWN
        else:
            result_response = {
                "response": "T√¥i ch∆∞a hi·ªÉu r√µ √Ω b·∫°n. H√£y th·ª≠ h·ªèi v·ªÅ s·∫£n ph·∫©m ho·∫∑c v·∫≠t li·ªáu!\n\n"
                           "**V√≠ d·ª•:**\n"
                           "‚Ä¢ \"T√¨m b√†n ƒÉn tr√≤n\"\n"
                           "‚Ä¢ \"T√¨m g·ªó s·ªìi\"\n"
                           "‚Ä¢ \"T√≠nh chi ph√≠ s·∫£n ph·∫©m B001\"\n"
                           "‚Ä¢ \"Xem v·∫≠t li·ªáu c·ªßa gh·∫ø G002\"",
                "suggested_prompts": [
                    "üîç T√¨m s·∫£n ph·∫©m",
                    "üß± T√¨m v·∫≠t li·ªáu",
                    "üìã Danh s√°ch nh√≥m v·∫≠t li·ªáu"
                ]
            }
        
        # L·∫•y th√¥ng tin m·ªü r·ªông t·ª´ k·∫øt qu·∫£ t√¨m ki·∫øm
        expanded = None
        keywords = []
        
        if intent == "search_product" and result_response.get("data"):
            expanded = result_response["data"].get("expanded_query")
            # L·∫•y keywords t·ª´ params
            if params.get("keywords_vector"):
                keywords = extract_product_keywords(params["keywords_vector"])
        
        save_chat_history(
            msg.session_id,
            user_message,
            result_response.get("response", ""),
            intent,
            params,
            result_count,
            search_type="text",
            expanded_query=expanded,
            extracted_keywords=keywords
        )


        if result_response:
            result_response["query"] = user_message 

        return result_response
    
    except Exception as e:
        print(f"Server Error: {e}")
        import traceback
        traceback.print_exc()
        return {"response": f"‚ö†Ô∏è L·ªói h·ªá th·ªëng: {str(e)}"}

# ========================================
# BATCH OPERATIONS FOR MULTIPLE PRODUCTS
# ========================================

class BatchProductRequest(BaseModel):
    product_headcodes: List[str]
    session_id: str = ""
    operation: str  # "detail", "materials", "cost"

@app.post("/batch/products")
def batch_product_operations(request: BatchProductRequest):
    """
    üî• X·ª≠ l√Ω batch operations cho nhi·ªÅu s·∫£n ph·∫©m
    Operations: detail, materials, cost
    """
    try:
        if not request.product_headcodes:
            return {"response": "‚ö†Ô∏è Vui l√≤ng ch·ªçn √≠t nh·∫•t 1 s·∫£n ph·∫©m"}
        
        headcodes = request.product_headcodes
        operation = request.operation
        
        print(f"üì¶ Batch {operation}: {len(headcodes)} products")
        
        # ========== OPERATION: CHI TI·∫æT S·∫¢N PH·∫®M ==========
        if operation == "detail":
            conn = get_db()
            cur = conn.cursor(cursor_factory=RealDictCursor)
            
            cur.execute("""
                SELECT headcode, product_name, category, sub_category, 
                       material_primary, project, unit
                FROM products
                WHERE headcode = ANY(%s)
                ORDER BY product_name
            """, (headcodes,))
            
            products = cur.fetchall()
            conn.close()
            
            if not products:
                return {"response": "‚ùå Kh√¥ng t√¨m th·∫•y s·∫£n ph·∫©m"}
            
            response = f"üìã **CHI TI·∫æT {len(products)} S·∫¢N PH·∫®M:**\n\n"
            
            for idx, prod in enumerate(products, 1):
                response += f"**{idx}. {prod['product_name']}**\n"
                response += f"   ‚Ä¢ M√£: `{prod['headcode']}`\n"
                response += f"   ‚Ä¢ Danh m·ª•c: {prod.get('category', 'N/A')}"
                
                if prod.get('sub_category'):
                    response += f" - {prod['sub_category']}"
                
                response += f"\n   ‚Ä¢ V·∫≠t li·ªáu ch√≠nh: {prod.get('material_primary', 'N/A')}\n"
                
                if prod.get('project'):
                    response += f"   ‚Ä¢ D·ª± √°n: {prod['project']}\n"
                
                response += "\n"
            
            return {
                "response": response,
                "products": [dict(p) for p in products]
            }
        
        # ========== OPERATION: ƒê·ªäNH M·ª®C V·∫¨T LI·ªÜU ==========
        elif operation == "materials":
            conn = get_db()
            cur = conn.cursor(cursor_factory=RealDictCursor)
            
            # L·∫•y t·∫•t c·∫£ v·∫≠t li·ªáu c·ªßa c√°c s·∫£n ph·∫©m
            cur.execute("""
                SELECT 
                    p.headcode,
                    p.product_name,
                    m.id_sap,
                    m.material_name,
                    m.material_group,
                    m.material_subprice,
                    m.unit,
                    pm.quantity,
                    pm.unit as pm_unit
                FROM product_materials pm
                INNER JOIN products p ON pm.product_headcode = p.headcode
                INNER JOIN materials m ON pm.material_id_sap = m.id_sap
                WHERE p.headcode = ANY(%s)
                ORDER BY p.product_name, m.material_name
            """, (headcodes,))
            
            records = cur.fetchall()
            conn.close()
            
            if not records:
                return {"response": "‚ö†Ô∏è C√°c s·∫£n ph·∫©m n√†y ch∆∞a c√≥ ƒë·ªãnh m·ª©c v·∫≠t li·ªáu"}
            
            # Group by product
            products_dict = {}
            for rec in records:
                hc = rec['headcode']
                if hc not in products_dict:
                    products_dict[hc] = {
                        'headcode': hc,
                        'product_name': rec['product_name'],
                        'materials': []
                    }
                
                price = get_latest_material_price(rec['material_subprice'])
                qty = float(rec['quantity']) if rec['quantity'] else 0.0
                
                products_dict[hc]['materials'].append({
                    'id_sap': rec['id_sap'],
                    'name': rec['material_name'],
                    'group': rec['material_group'],
                    'quantity': qty,
                    'unit': rec['pm_unit'],
                    'price': price,
                    'total': qty * price
                })
            
            # T·∫°o response
            response = f"üß± **ƒê·ªäNH M·ª®C V·∫¨T LI·ªÜU - {len(products_dict)} S·∫¢N PH·∫®M:**\n\n"
            
            for prod_data in products_dict.values():
                response += f"### üì¶ {prod_data['product_name']} (`{prod_data['headcode']}`)\n\n"
                
                total_cost = sum(m['total'] for m in prod_data['materials'])
                
                for idx, mat in enumerate(prod_data['materials'][:10], 1):
                    response += f"{idx}. **{mat['name']}** ({mat['group']})\n"
                    response += f"   ‚Ä¢ S·ªë l∆∞·ª£ng: {mat['quantity']} {mat['unit']}\n"
                    response += f"   ‚Ä¢ ƒê∆°n gi√°: {mat['price']:,.0f} VNƒê\n"
                    response += f"   ‚Ä¢ Th√†nh ti·ªÅn: **{mat['total']:,.0f} VNƒê**\n\n"
                
                if len(prod_data['materials']) > 10:
                    response += f"*...v√† {len(prod_data['materials'])-10} v·∫≠t li·ªáu kh√°c*\n\n"
                
                response += f"üí∞ **T·ªïng NVL ({prod_data['headcode']}): {total_cost:,.0f} VNƒê**\n\n"
                response += "---\n\n"
            
            # T·∫°o materials list ƒë·ªÉ UI c√≥ th·ªÉ render cards
            all_materials = []
            for prod_data in products_dict.values():
                all_materials.extend(prod_data['materials'])
            
            return {
                "response": response,
                "products_materials": products_dict,
                "materials": all_materials  # ƒê·ªÉ UI render material cards
            }
        
        # ========== OPERATION: CHI PH√ç ==========
        elif operation == "cost":
            conn = get_db()
            cur = conn.cursor(cursor_factory=RealDictCursor)
            
            cur.execute("""
                SELECT 
                    p.headcode,
                    p.product_name,
                    p.category,
                    m.material_name,
                    m.material_group,
                    m.material_subprice,
                    pm.quantity,
                    pm.unit
                FROM product_materials pm
                INNER JOIN products p ON pm.product_headcode = p.headcode
                INNER JOIN materials m ON pm.material_id_sap = m.id_sap
                WHERE p.headcode = ANY(%s)
                ORDER BY p.product_name
            """, (headcodes,))
            
            records = cur.fetchall()
            conn.close()
            
            if not records:
                return {"response": "‚ö†Ô∏è Kh√¥ng c√≥ d·ªØ li·ªáu ƒë·ªãnh m·ª©c"}
            
            # T√≠nh chi ph√≠ t·ª´ng s·∫£n ph·∫©m
            products_cost = {}
            for rec in records:
                hc = rec['headcode']
                if hc not in products_cost:
                    products_cost[hc] = {
                        'headcode': hc,
                        'name': rec['product_name'],
                        'category': rec['category'],
                        'material_cost': 0.0,
                        'materials_detail': []
                    }
                
                qty = float(rec['quantity']) if rec['quantity'] else 0.0
                price = get_latest_material_price(rec['material_subprice'])
                total = qty * price
                
                products_cost[hc]['material_cost'] += total
                products_cost[hc]['materials_detail'].append({
                    'name': rec['material_name'],
                    'group': rec['material_group'],
                    'quantity': qty,
                    'unit': rec['unit'],
                    'price': price,
                    'total': total
                })
            
            # Response
            response = f"üí∞ **B√ÅO C√ÅO CHI PH√ç - {len(products_cost)} S·∫¢N PH·∫®M:**\n\n"
            
            grand_total = 0.0
            
            for prod_data in products_cost.values():
                response += f"### üì¶ {prod_data['name']} (`{prod_data['headcode']}`)\n"
                response += f"**Danh m·ª•c:** {prod_data['category']}\n\n"
                response += f"**Chi ph√≠ nguy√™n v·∫≠t li·ªáu:** {prod_data['material_cost']:,.0f} VNƒê\n"
                response += f"   ‚Ä¢ {len(prod_data['materials_detail'])} lo·∫°i v·∫≠t li·ªáu\n\n"
                response += "---\n\n"
                
                grand_total += prod_data['material_cost']
            
            response += f"## üíµ T·ªîNG CHI PH√ç NVL: {grand_total:,.0f} VNƒê\n\n"
            response += "üìã *Chi ph√≠ ƒë∆∞·ª£c t√≠nh t·ª´ gi√° nguy√™n v·∫≠t li·ªáu g·∫ßn nh·∫•t*"
            
            return {
                "response": response,
                "products_cost": products_cost,
                "grand_total": grand_total
            }
        
        else:
            return {"response": "‚ùå Operation kh√¥ng h·ª£p l·ªá"}
    
    except Exception as e:
        print(f"‚ùå Batch operation error: {e}")
        import traceback
        traceback.print_exc()
        return {"response": f"‚ùå L·ªói: {str(e)}"}
# ========================================
# MODULE 1: CONSOLIDATED BOM REPORT
# ========================================

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class ConsolidatedBOMRequest(BaseModel):
    product_headcodes: List[str]
    session_id: str = ""

import re  # <--- Th√™m c√°i n√†y
from io import BytesIO
from typing import List
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# --- T·ª∞ ƒê·ªäNH NGHƒ®A REGEX ƒê·ªÇ L·ªåC K√ù T·ª∞ L·ªñI ---
# Regex n√†y l·ªçc c√°c k√Ω t·ª± ASCII ƒëi·ªÅu khi·ªÉn (Control chars) kh√¥ng h·ª£p l·ªá trong file Excel (XML)
# Bao g·ªìm: ASCII 0-8, 11-12, 14-31
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def generate_consolidated_report(product_headcodes: List[str]) -> BytesIO:
    """
    T·∫°o b√°o c√°o Excel t·ªïng h·ª£p ƒë·ªãnh m·ª©c v·∫≠t t∆∞ cho nhi·ªÅu s·∫£n ph·∫©m
    
    Args:
        product_headcodes: Danh s√°ch m√£ s·∫£n ph·∫©m
    
    Returns:
        BytesIO: File Excel buffer
    """
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    # 1. L·∫§Y TH√îNG TIN S·∫¢N PH·∫®M
    cur.execute("""
        SELECT headcode, product_name, category, sub_category, project
        FROM products 
        WHERE headcode = ANY(%s)
        ORDER BY product_name
    """, (product_headcodes,))
    
    selected_products = cur.fetchall()
    
    if not selected_products:
        raise ValueError("Kh√¥ng t√¨m th·∫•y s·∫£n ph·∫©m n√†o")
    
    # 2. L·∫§Y ƒê·ªäNH M·ª®C CHI TI·∫æT (Flatten View)
    cur.execute("""
        SELECT 
            p.headcode,
            p.product_name,
            m.id_sap,
            m.material_name,
            m.material_group,
            m.material_subgroup,
            m.unit as material_unit,
            pm.quantity,
            pm.unit as pm_unit,
            m.material_subprice
        FROM product_materials pm
        INNER JOIN products p ON pm.product_headcode = p.headcode
        INNER JOIN materials m ON pm.material_id_sap = m.id_sap
        WHERE p.headcode = ANY(%s)
        ORDER BY p.product_name, m.material_name
    """, (product_headcodes,))
    
    detail_records = cur.fetchall()
    conn.close()
    
    if not detail_records:
        raise ValueError("C√°c s·∫£n ph·∫©m n√†y ch∆∞a c√≥ ƒë·ªãnh m·ª©c v·∫≠t t∆∞")
    
    # 3. AGGREGATION - G·ªòP V·∫¨T T∆Ø
    material_summary = {}
    
    for record in detail_records:
        id_sap = record['id_sap']
        quantity = float(record['quantity']) if record['quantity'] else 0.0
        
        # Parse gi√° m·ªõi nh·∫•t
        latest_price = get_latest_material_price(record['material_subprice'])
        
        if id_sap not in material_summary:
            material_summary[id_sap] = {
                'id_sap': id_sap,
                'material_name': record['material_name'],
                'material_group': record['material_group'],
                'material_subgroup': record['material_subgroup'],
                'unit': record['material_unit'],
                'total_quantity': 0.0,
                'unit_price': latest_price,
                'total_cost': 0.0,
                'used_in_products': []
            }
        
        # C·ªông d·ªìn s·ªë l∆∞·ª£ng
        material_summary[id_sap]['total_quantity'] += quantity
        material_summary[id_sap]['used_in_products'].append(
            f"{record['product_name']} ({quantity} {record['pm_unit']})"
        )
    
    # T√≠nh th√†nh ti·ªÅn
    for mat_id, mat_data in material_summary.items():
        mat_data['total_cost'] = mat_data['total_quantity'] * mat_data['unit_price']
    
    # 4. T·∫†O EXCEL FILE
    wb = Workbook()
    
    # --- SHEET 1: OVERVIEW (Danh s√°ch SP ƒë√£ ch·ªçn) ---
    ws_overview = wb.active
    ws_overview.title = "Overview"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Headers
    overview_headers = ["STT", "M√£ SP", "T√™n S·∫£n Ph·∫©m", "Danh M·ª•c", "D·ª± √Ån"]
    for col_idx, header in enumerate(overview_headers, 1):
        cell = ws_overview.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for idx, prod in enumerate(selected_products, 1):
        ws_overview.append([
            idx,
            prod['headcode'],
            prod['product_name'],
            f"{prod.get('category', '')} - {prod.get('sub_category', '')}",
            prod.get('project', '')
        ])
    
    # Auto-adjust column width
    for col in ws_overview.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws_overview.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    # --- SHEET 2: MATERIAL SUMMARY (T·ªïng h·ª£p v·∫≠t t∆∞) ---
    ws_summary = wb.create_sheet("Material Summary")
    
    summary_headers = [
        "STT", "M√£ SAP", "T√™n V·∫≠t Li·ªáu", "Nh√≥m", 
        "Nh√≥m Con", "ƒê∆°n V·ªã", "T·ªïng SL", "ƒê∆°n Gi√° (VNƒê)", "Th√†nh Ti·ªÅn (VNƒê)"
    ]
    
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Sort by total_cost DESC
    sorted_materials = sorted(
        material_summary.values(), 
        key=lambda x: x['total_cost'], 
        reverse=True
    )
    
    total_cost_all = 0.0
    
    for idx, mat in enumerate(sorted_materials, 1):
        ws_summary.append([
            idx,
            mat['id_sap'],
            mat['material_name'],
            mat['material_group'],
            mat['material_subgroup'],
            mat['unit'],
            round(mat['total_quantity'], 2),
            round(mat['unit_price'], 2),
            round(mat['total_cost'], 2)
        ])
        total_cost_all += mat['total_cost']
    
    # T·ªîNG C·ªòNG ROW
    summary_row = ws_summary.max_row + 1
    ws_summary.cell(row=summary_row, column=7, value="T·ªîNG C·ªòNG:").font = Font(bold=True)
    ws_summary.cell(row=summary_row, column=9, value=round(total_cost_all, 2)).font = Font(bold=True, color="FF0000")
    
    for col in ws_summary.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws_summary.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    # --- SHEET 3: DETAILS (Chi ti·∫øt theo SP) ---
    ws_details = wb.create_sheet("Details")
    
    detail_headers = [
        "M√£ SP", "T√™n SP", "M√£ SAP", "T√™n V·∫≠t Li·ªáu", 
        "Nh√≥m VL", "S·ªë L∆∞·ª£ng", "ƒê∆°n V·ªã", "ƒê∆°n Gi√°", "Th√†nh Ti·ªÅn"
    ]
    
    for col_idx, header in enumerate(detail_headers, 1):
        cell = ws_details.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for record in detail_records:
        quantity = float(record['quantity']) if record['quantity'] else 0.0
        unit_price = get_latest_material_price(record['material_subprice'])
        total_cost = quantity * unit_price
        
        ws_details.append([
            record['headcode'],
            record['product_name'],
            record['id_sap'],
            record['material_name'],
            record['material_group'],
            round(quantity, 2),
            record['pm_unit'],
            round(unit_price, 2),
            round(total_cost, 2)
        ])
    
    for col in ws_details.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws_details.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    # 5. SAVE TO BUFFER
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


@app.post("/report/consolidated")
def create_consolidated_report(request: ConsolidatedBOMRequest):
    """
    üìä API Endpoint t·∫°o b√°o c√°o t·ªïng h·ª£p ƒë·ªãnh m·ª©c v·∫≠t t∆∞
    
    Input: {"product_headcodes": ["B001", "B002", "G001"], "session_id": "..."}
    Output: File Excel (.xlsx)
    """
    try:
        if not request.product_headcodes or len(request.product_headcodes) == 0:
            return {"message": "‚ö†Ô∏è Vui l√≤ng ch·ªçn √≠t nh·∫•t 1 s·∫£n ph·∫©m"}
        
        print(f"üìä Generating report for {len(request.product_headcodes)} products...")
        
        # T·∫°o file Excel
        excel_buffer = generate_consolidated_report(request.product_headcodes)
        
        # L∆∞u l·ªãch s·ª≠ (Optional)
        if request.session_id:
            save_chat_history(
                session_id=request.session_id,
                user_message=f"[REPORT] T·ªïng h·ª£p {len(request.product_headcodes)} s·∫£n ph·∫©m",
                bot_response="ƒê√£ t·∫°o b√°o c√°o Excel",
                intent="generate_report",
                params={"products": request.product_headcodes},
                result_count=len(request.product_headcodes),
                search_type="report"
            )
        
        # Tr·∫£ file v·ªÅ client
        from fastapi.responses import StreamingResponse
        
        filename = f"BOM_Consolidated_{len(request.product_headcodes)}SP_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except ValueError as e:
        return {"message": f"‚ùå {str(e)}"}
    except Exception as e:
        print(f"‚ùå Report generation error: {e}")
        import traceback
        traceback.print_exc()
        return {"message": f"‚ùå L·ªói t·∫°o b√°o c√°o: {str(e)}"}


# ========================================
# MODULE 2: ADAPTIVE USER LEARNING
# ========================================

import numpy as np

class TrackingRequest(BaseModel):
    session_id: str
    product_headcode: str
    interaction_type: str  # 'view', 'reject', 'select'

@app.post("/track/view")
def track_product_view(request: TrackingRequest):
    """
    üëÅÔ∏è Track khi user XEM CHI TI·∫æT s·∫£n ph·∫©m (Positive Signal)
    """
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # L·∫•y embedding c·ªßa s·∫£n ph·∫©m
        cur.execute("""
            SELECT description_embedding 
            FROM products 
            WHERE headcode = %s AND description_embedding IS NOT NULL
        """, (request.product_headcode,))
        
        result = cur.fetchone()
        
        if not result:
            conn.close()
            return {"message": "Product not found or no embedding"}
        
        product_vector = result['description_embedding']
        
        # L∆∞u v√†o user_preferences
        cur.execute("""
            INSERT INTO user_preferences 
            (session_id, product_headcode, product_vector, interaction_type, weight)
            VALUES (%s, %s, %s, %s, %s)
        """, (
            request.session_id,
            request.product_headcode,
            product_vector,
            'view',
            1.0  # Positive signal
        ))
        
        conn.commit()
        conn.close()
        
        print(f"‚úÖ Tracked VIEW: {request.product_headcode} by {request.session_id[:8]}")
        
        return {"message": "‚úÖ Tracked successfully", "type": "view"}
        
    except Exception as e:
        print(f"‚ùå Tracking error: {e}")
        return {"message": f"Error: {str(e)}"}


@app.post("/track/reject")
def track_product_reject(request: TrackingRequest):
    """
    ‚ùå Track khi user B·ªé QUA/REJECT s·∫£n ph·∫©m (Negative Signal)
    """
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT description_embedding 
            FROM products 
            WHERE headcode = %s AND description_embedding IS NOT NULL
        """, (request.product_headcode,))
        
        result = cur.fetchone()
        
        if not result:
            conn.close()
            return {"message": "Product not found"}
        
        product_vector = result['description_embedding']
        
        cur.execute("""
            INSERT INTO user_preferences 
            (session_id, product_headcode, product_vector, interaction_type, weight)
            VALUES (%s, %s, %s, %s, %s)
        """, (
            request.session_id,
            request.product_headcode,
            product_vector,
            'reject',
            -1.0  # Negative signal
        ))
        
        conn.commit()
        conn.close()
        
        print(f"‚ùå Tracked REJECT: {request.product_headcode} by {request.session_id[:8]}")
        
        return {"message": "‚úÖ Tracked rejection", "type": "reject"}
        
    except Exception as e:
        print(f"‚ùå Tracking error: {e}")
        return {"message": f"Error: {str(e)}"}


# PERSONALIZATION - REFACTORED V5.7
# ========================================

def calculate_personalized_score(
    candidate_vector: list, 
    session_id: str
) -> float:
    """
    üéØ V5.7 - Tr·∫£ v·ªÅ ƒëi·ªÉm Personalization RI√äNG (0.0 ‚Üí 1.0)
    KH√îNG tr·∫£ v·ªÅ final_score, ƒë·ªÉ search_products t·ªïng h·ª£p sau
    
    Returns:
        float: Personal affinity score (0.0 = kh√¥ng kh·ªõp, 1.0 = r·∫•t kh·ªõp)
    """
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # L·∫•y 10 interactions g·∫ßn nh·∫•t
        cur.execute("""
            SELECT product_vector, weight
            FROM user_preferences
            WHERE session_id = %s
            ORDER BY created_at DESC
            LIMIT 10
        """, (session_id,))
        
        history = cur.fetchall()
        conn.close()
        
        if not history:
            return 0.5  # Neutral score khi ch∆∞a c√≥ history
        
        # Convert candidate sang numpy
        if isinstance(candidate_vector, str):
            candidate_np = np.array(json.loads(candidate_vector), dtype=np.float32)
        else:
            candidate_np = np.array(candidate_vector, dtype=np.float32)
        
        positive_scores = []
        negative_scores = []
        
        for record in history:
            try:
                vec_data = record['product_vector']
                
                if vec_data is None:
                    continue
                    
                # Parse vector
                if isinstance(vec_data, str):
                    hist_vector = np.array(json.loads(vec_data), dtype=np.float32)
                elif isinstance(vec_data, list):
                    hist_vector = np.array(vec_data, dtype=np.float32)
                else:
                    continue
                
                # Check dimension match
                if len(hist_vector) != len(candidate_np):
                    continue
                
                # Cosine Similarity
                norm_product = np.linalg.norm(candidate_np) * np.linalg.norm(hist_vector)
                if norm_product < 1e-8:
                    continue
                    
                similarity = np.dot(candidate_np, hist_vector) / norm_product
                
                # Ph√¢n lo·∫°i theo weight
                if record['weight'] > 0:
                    positive_scores.append(similarity)
                else:
                    negative_scores.append(similarity)
                    
            except Exception:
                continue
        
        # Fallback n·∫øu kh√¥ng c√≥ scores h·ª£p l·ªá
        if not positive_scores and not negative_scores:
            return 0.5
        
        # T√≠nh ƒëi·ªÉm affinity thu·∫ßn t√∫y
        positive_affinity = np.mean(positive_scores) if positive_scores else 0.0
        negative_penalty = np.mean(negative_scores) if negative_scores else 0.0
        
        # Formula: Positive boost - Negative penalty
        personal_score = positive_affinity - (negative_penalty * 0.5)
        
        # Clip v·ªÅ [0, 1]
        personal_score = float(np.clip(personal_score, 0.0, 1.0))
        
        return personal_score
        
    except Exception as e:
        print(f"‚ö†Ô∏è Personalization error: {e}")
        return 0.5
# ========================================
# NEW ENDPOINT: USER FEEDBACK
# ========================================

class FeedbackRequest(BaseModel):
    session_id: str
    query: str
    selected_items: List[str]  # List of headcodes ho·∫∑c id_sap
    rejected_items: List[str] = []
    search_type: str  # "product" ho·∫∑c "material"

@app.post("/feedback")
def submit_feedback(feedback: FeedbackRequest):
    """
    üìù Endpoint nh·∫≠n feedback t·ª´ user v·ªÅ k·∫øt qu·∫£ t√¨m ki·∫øm
    """
    try:
        success = save_user_feedback(
            feedback.session_id,
            feedback.query,
            feedback.selected_items,
            feedback.rejected_items,
            feedback.search_type
        )
        
        if success:
            return {
                "message": "‚úÖ C·∫£m ∆°n ph·∫£n h·ªìi c·ªßa b·∫°n! K·∫øt qu·∫£ t√¨m ki·∫øm s·∫Ω ƒë∆∞·ª£c c·∫£i thi·ªán.",
                "saved": True
            }
        else:
            return {
                "message": "‚ö†Ô∏è Kh√¥ng th·ªÉ l∆∞u ph·∫£n h·ªìi",
                "saved": False
            }
            
    except Exception as e:
        return {
            "message": f"‚ùå L·ªói: {str(e)}",
            "saved": False
        }




# ========================================
# IMAGE SEARCH
# ========================================

@app.post("/search-image")
async def search_by_image(
    file: UploadFile = File(...),
    session_id: str = Form(default=str(uuid.uuid4()))
):
    """T√¨m ki·∫øm theo ·∫£nh"""
    file_path = f"temp_{uuid.uuid4()}.jpg"
    try:
        with open(file_path, "wb") as buffer:
            import shutil
            shutil.copyfileobj(file.file, buffer)
        
        img = Image.open(file_path)
        model = genai.GenerativeModel("gemini-2.5-flash")
        
        prompt = """
        ƒê√≥ng vai chuy√™n gia k·ªπ thu·∫≠t AA Corporation.
        Ph√¢n t√≠ch ·∫£nh n·ªôi th·∫•t n√†y ƒë·ªÉ tr√≠ch xu·∫•t th√¥ng tin t√¨m ki·∫øm Database.
        
        OUTPUT JSON ONLY (no markdown, no backticks):
        {
          "category": "Lo·∫°i SP (B√†n, Gh·∫ø, Sofa...)",
          "visual_description": "M√¥ t·∫£ chi ti·∫øt k·ªπ thu·∫≠t d√πng cho Vector Search",
          "material_detected": "V·∫≠t li·ªáu ch√≠nh (G·ªó, Da, V·∫£i, ƒê√°...)",
          "color_tone": "M√†u ch·ªß ƒë·∫°o"
        }
        """
        
        response = model.generate_content([prompt, img])
        
        if not response.text:
            return {
                "response": "‚ö†Ô∏è Kh√¥ng ph√¢n t√≠ch ƒë∆∞·ª£c ·∫£nh. Vui l√≤ng th·ª≠ ·∫£nh kh√°c.",
                "products": []
            }
        
        clean = response.text.strip()
        
        if "```json" in clean:
            clean = clean.split("```json")[1].split("```")[0].strip()
        elif "```" in clean:
            clean = clean.split("```")[1].split("```")[0].strip()
        
        try:
            ai_result = json.loads(clean)
        except json.JSONDecodeError as e:
            print(f"JSON Parse Error: {e}")
            ai_result = {
                "visual_description": clean[:200],
                "category": "N·ªôi th·∫•t"
            }
        
        params = {
            "category": ai_result.get("category"),
            "keywords_vector": ai_result.get("visual_description"),
            "material_primary": ai_result.get("material_detected")
        }
        
        search_result = search_products(params)
        products = search_result.get("products", [])
        
        save_chat_history(
            session_id=session_id,
            user_message="[IMAGE_UPLOAD]",
            bot_response=f"Ph√¢n t√≠ch ·∫£nh: {ai_result.get('visual_description', 'N/A')[:100]}... | T√¨m th·∫•y {len(products)} s·∫£n ph·∫©m",
            intent="search_product",
            params=params,
            result_count=len(products),
            search_type="image",
            expanded_query=ai_result.get("visual_description"),
            extracted_keywords=[
                ai_result.get("category"),
                ai_result.get("material_detected"),
                ai_result.get("color_tone")
            ]
        )

        if not products:
            return {
                "response": f"üìã **Ph√¢n t√≠ch ·∫£nh:** T√¥i nh·∫≠n th·∫•y ƒë√¢y l√† **{ai_result.get('visual_description', 's·∫£n ph·∫©m n·ªôi th·∫•t')}**.\n\n"
                                f"Tuy nhi√™n, kh√¥ng t√¨m th·∫•y s·∫£n ph·∫©m t∆∞∆°ng t·ª± trong kho d·ªØ li·ªáu.\n\n"
                        f"üí° **G·ª£i √Ω**: B·∫°n c√≥ th·ªÉ m√¥ t·∫£ chi ti·∫øt h∆°n. Ho·∫∑c b·∫°n c√≥ th·ªÉ t√¨m s·∫£n ph·∫©m kh√°c. T√¥i s·∫Ω g·ª£i √Ω cho b·∫°n danh s√°ch s·∫£n ph·∫©m",
                "products": [],
                "ai_interpretation": ai_result.get("visual_description", "")
            }
        
        return {
            "response": f"üìã **Ph√¢n t√≠ch ·∫£nh:** T√¥i nh·∫≠n th·∫•y ƒë√¢y l√† **{ai_result.get('visual_description', 's·∫£n ph·∫©m')}**.\n\n"
                       f"‚úÖ ƒê√£ t√¨m th·∫•y **{len(products)} s·∫£n ph·∫©m** t∆∞∆°ng ƒë·ªìng:",
            "products": products,
            "ai_interpretation": ai_result.get("visual_description", ""),
            "search_method": "image_vector"
        }
    
    except Exception as e:
        print(f"‚ùå Image search error: {e}")
        import traceback
        traceback.print_exc()
        
        return {
            "response": f"‚ö†Ô∏è L·ªói x·ª≠ l√Ω ·∫£nh: {str(e)}. Vui l√≤ng th·ª≠ l·∫°i.",
            "products": []
        }
    
    finally:
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except:
                pass

# ========================================
# IMPORT ENDPOINTS
# ========================================
# ========================================
# TH√äM V√ÄO chatbot_api.py
# ========================================

# [1] BATCH CLASSIFICATION FUNCTIONS
# Th√™m sau ph·∫ßn AUTO CLASSIFICATION AI (d√≤ng ~100)

def batch_classify_products(products_batch: List[Dict]) -> List[Dict]:
    """
    Ph√¢n lo·∫°i H√ÄNG LO·∫†T s·∫£n ph·∫©m - 1 API call cho nhi·ªÅu s·∫£n ph·∫©m
    Input: [{'name': 'B√ÄN G·ªñ', 'id_sap': 'SP001'}, ...]
    Output: [{'id_sap': 'SP001', 'category': 'B√†n', ...}, ...]
    """
    if not products_batch:
        return []
    
    # [FIX] ƒê·ªïi sang model ·ªïn ƒë·ªãnh ƒë·ªÉ tr√°nh l·ªói Rate Limit c·ªßa b·∫£n Experimental
    model = genai.GenerativeModel("gemini-2.5-flash")
    
    # T·∫°o danh s√°ch s·∫£n ph·∫©m trong prompt
    products_text = ""
    for i, prod in enumerate(products_batch, 1):
        products_text += f"{i}. ID: {prod['id_sap']}, T√™n: {prod['name']}\n"
    
    prompt = f"""
B·∫°n l√† chuy√™n gia ph√¢n lo·∫°i s·∫£n ph·∫©m n·ªôi th·∫•t cao c·∫•p.

Ph√¢n lo·∫°i {len(products_batch)} s·∫£n ph·∫©m sau:

{products_text}

M·ªói s·∫£n ph·∫©m c·∫ßn ph√¢n lo·∫°i theo:
1. category: B√†n, Gh·∫ø, Sofa, T·ªß, Gi∆∞·ªùng, ƒê√®n, K·ªá, B√†n l√†m vi·ªác, Kh√°c
2. sub_category: Danh m·ª•c ph·ª• c·ª• th·ªÉ (VD: "B√†n ƒÉn", "Gh·∫ø bar", "Sofa g√≥c"...)
3. material_primary: G·ªó, Da, V·∫£i, Kim lo·∫°i, ƒê√°, K√≠nh, Nh·ª±a, M√¢y tre, H·ªón h·ª£p

OUTPUT JSON ARRAY ONLY (no markdown, no backticks):
[
  {{"id_sap": "SP001", "category": "...", "sub_category": "...", "material_primary": "..."}},
  {{"id_sap": "SP002", "category": "...", "sub_category": "...", "material_primary": "..."}}
]
"""
    
    # G·ªçi AI v·ªõi retry logic
    response_text = call_gemini_with_retry(model, prompt, max_retries=3)
    
    # Fallback m·∫∑c ƒë·ªãnh n·∫øu AI l·ªói h·∫≥n
    default_results = [{
        'id_sap': p['id_sap'],
        'category': 'Ch∆∞a ph√¢n lo·∫°i',
        'sub_category': 'Ch∆∞a ph√¢n lo·∫°i',
        'material_primary': 'Ch∆∞a x√°c ƒë·ªãnh'
    } for p in products_batch]

    if not response_text:
        return default_results
    
    try:
        clean = response_text.strip()
        # X·ª≠ l√Ω tr∆∞·ªùng h·ª£p Gemini tr·∫£ v·ªÅ markdown code block
        if "```json" in clean:
            clean = clean.split("```json")[1].split("```")[0].strip()
        elif "```" in clean:
            clean = clean.split("```")[1].split("```")[0].strip()
        
        results = json.loads(clean)
        
        # ƒê·∫£m b·∫£o s·ªë l∆∞·ª£ng k·∫øt qu·∫£ kh·ªõp v·ªõi input
        if len(results) != len(products_batch):
            print(f"‚ö†Ô∏è Batch size mismatch: expected {len(products_batch)}, got {len(results)}")
            return default_results
        
        return results
        
    except Exception as e:
        print(f"‚ùå Batch classification parse error: {e}")
        return default_results
def batch_classify_materials(materials_batch: List[Dict]) -> List[Dict]:
    """
    Ph√¢n lo·∫°i H√ÄNG LO·∫†T v·∫≠t li·ªáu
    Input: [{'name': 'G·ªñ S·ªíI', 'id_sap': 'M001'}, ...]
    Output: [{'id_sap': 'M001', 'material_group': 'G·ªó', ...}, ...]
    """
    if not materials_batch:
        return []
    
    # [FIX] ƒê·ªïi sang model gemini-1.5-flash ƒë·ªÉ ·ªïn ƒë·ªãnh h∆°n v√† tr√°nh l·ªói Rate Limit
    model = genai.GenerativeModel("gemini-2.5-flash")
    
    materials_text = ""
    for i, mat in enumerate(materials_batch, 1):
        materials_text += f"{i}. ID: {mat['id_sap']}, T√™n: {mat['name']}\n"
    
    prompt = f"""
        Ph√¢n lo·∫°i {len(materials_batch)} nguy√™n v·∫≠t li·ªáu n·ªôi th·∫•t:

        {materials_text}

        X√°c ƒë·ªãnh:
        1. material_group: G·ªó, Da, V·∫£i, ƒê√°, Kim lo·∫°i, K√≠nh, Nh·ª±a, S∆°n, Keo, Ph·ª• ki·ªán, Kh√°c
        2. material_subgroup: Nh√≥m con c·ª• th·ªÉ (VD: "G·ªó t·ª± nhi√™n", "Da th·∫≠t", "V·∫£i cao c·∫•p")

        OUTPUT JSON ARRAY ONLY:
        [
        {{"id_sap": "M001", "material_group": "...", "material_subgroup": "..."}},
        {{"id_sap": "M002", "material_group": "...", "material_subgroup": "..."}}
        ]
    """
    
    # G·ªçi Gemini v·ªõi retry
    response_text = call_gemini_with_retry(model, prompt, max_retries=3)
    
    # T·∫°o k·∫øt qu·∫£ m·∫∑c ƒë·ªãnh (Fallback) ƒë·ªÉ tr·∫£ v·ªÅ n·∫øu AI l·ªói
    default_results = [{
        'id_sap': m['id_sap'],
        'material_group': 'Ch∆∞a ph√¢n lo·∫°i',
        'material_subgroup': 'Ch∆∞a ph√¢n lo·∫°i'
    } for m in materials_batch]

    if not response_text:
        return default_results
    
    try:
        clean = response_text.strip()
        # X·ª≠ l√Ω l√†m s·∫°ch markdown JSON
        if "```json" in clean:
            clean = clean.split("```json")[1].split("```")[0].strip()
        elif "```" in clean:
            clean = clean.split("```")[1].split("```")[0].strip()
        
        results = json.loads(clean)
        
        # Ki·ªÉm tra s·ªë l∆∞·ª£ng k·∫øt qu·∫£ tr·∫£ v·ªÅ c√≥ kh·ªõp input kh√¥ng
        if len(results) != len(materials_batch):
            print(f"‚ö†Ô∏è Batch materials mismatch: expected {len(materials_batch)}, got {len(results)}")
            return default_results
        
        return results
        
    except Exception as e:
        print(f"‚ùå Batch materials classification error: {e}")
        return default_results


# Thay th·∫ø 2 endpoints import c≈©
# ========================================

@app.post("/import/products")
async def import_products(file: UploadFile = File(...)):
    """
    [V4.1] Import products - KH√îNG auto classify ngay
    Ch·ªâ import v√†o DB, classify sau qua endpoint ri√™ng
    """
    try:
        contents = await file.read()
        df = pd.read_csv(io.BytesIO(contents))
        
        # Chu·∫©n h√≥a t√™n c·ªôt
        df.columns = df.columns.str.strip().str.lower()
        
        required = ['headcode', 'id_sap', 'product_name']
        missing = [col for col in required if col not in df.columns]
        
        if missing:
            return {
                "message": f"‚ùå Thi·∫øu c√°c c·ªôt b·∫Øt bu·ªôc: {', '.join(missing)}",
                "required_columns": required,
                "your_columns": list(df.columns)
            }
        
        conn = get_db()
        cur = conn.cursor()
        
        imported = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                headcode = str(row['headcode']).strip()
                id_sap = str(row['id_sap']).strip()
                product_name = str(row['product_name']).strip()
                
                if not headcode or not id_sap or not product_name:
                    errors.append(f"Row {idx+2}: Missing required fields")
                    continue
                
                # L·∫§Y TR·ª∞C TI·∫æP t·ª´ CSV (n·∫øu c√≥), KH√îNG g·ªçi AI
                category = str(row.get('category', 'Ch∆∞a ph√¢n lo·∫°i')).strip() if pd.notna(row.get('category')) else 'Ch∆∞a ph√¢n lo·∫°i'
                sub_category = str(row.get('sub_category', 'Ch∆∞a ph√¢n lo·∫°i')).strip() if pd.notna(row.get('sub_category')) else 'Ch∆∞a ph√¢n lo·∫°i'
                material_primary = str(row.get('material_primary', 'Ch∆∞a x√°c ƒë·ªãnh')).strip() if pd.notna(row.get('material_primary')) else 'Ch∆∞a x√°c ƒë·ªãnh'
                
                unit = str(row.get('unit', '')).strip() if pd.notna(row.get('unit')) else None
                project = str(row.get('project', '')).strip() if pd.notna(row.get('project')) else None
                project_id = str(row.get('project_id', '')).strip() if pd.notna(row.get('project_id')) else None
                
                sql = """
                    INSERT INTO products (
                        headcode, id_sap, product_name, 
                        category, sub_category, material_primary,
                        unit, project, project_id
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (headcode) DO UPDATE SET
                        product_name = EXCLUDED.product_name,
                        category = EXCLUDED.category,
                        sub_category = EXCLUDED.sub_category,
                        material_primary = EXCLUDED.material_primary,
                        unit = EXCLUDED.unit,
                        project = EXCLUDED.project,
                        project_id = EXCLUDED.project_id,
                        updated_at = NOW()
                """
                
                cur.execute(sql, (
                    headcode, id_sap, product_name,
                    category, sub_category, material_primary,
                    unit, project, project_id
                ))
                
                imported += 1
                
            except Exception as e:
                errors.append(f"Row {idx+2}: {str(e)[:100]}")
        
        conn.commit()
        conn.close()
        
        # ƒê·∫øm s·ªë s·∫£n ph·∫©m c·∫ßn classify
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM products 
            WHERE category = 'Ch∆∞a ph√¢n lo·∫°i' 
            OR sub_category = 'Ch∆∞a ph√¢n lo·∫°i'
            OR material_primary = 'Ch∆∞a x√°c ƒë·ªãnh'
        """)
        pending_count = cur.fetchone()[0]
        conn.close()
        
        message = f"‚úÖ Import th√†nh c√¥ng {imported}/{len(df)} products"
        if pending_count > 0:
            message += f"\n\n‚è≥ C√≥ {pending_count} s·∫£n ph·∫©m ch∆∞a ph√¢n lo·∫°i."
            message += f"\nüí° D√πng n√∫t 'ü§ñ Auto Classify' trong sidebar ƒë·ªÉ ph√¢n lo·∫°i h√†ng lo·∫°t."
        
        return {
            "message": message,
            "imported": imported,
            "total": len(df),
            "pending_classification": pending_count,
            "errors": errors[:10] if errors else []
        }
        
    except Exception as e:
        return {"message": f"‚ùå L·ªói: {str(e)}"}


@app.post("/import/materials")
async def import_materials(file: UploadFile = File(...)):
    """
    [V4.1] Import materials - KH√îNG auto classify ngay
    """
    try:
        contents = await file.read()
        df = pd.read_csv(io.BytesIO(contents))
        
        df.columns = df.columns.str.strip().str.lower()
        
        required = ['id_sap', 'material_name', 'material_group']
        missing = [col for col in required if col not in df.columns]
        
        if missing:
            return {
                "message": f"‚ùå Thi·∫øu c√°c c·ªôt b·∫Øt bu·ªôc: {', '.join(missing)}",
                "required_columns": required,
                "your_columns": list(df.columns)
            }
        
        conn = get_db()
        cur = conn.cursor()
        
        imported = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                id_sap = str(row['id_sap']).strip()
                material_name = str(row['material_name']).strip()
                material_group = str(row['material_group']).strip()
                
                if not id_sap or not material_name or not material_group:
                    errors.append(f"Row {idx+2}: Missing required fields")
                    continue
                
                # KH√îNG g·ªçi AI ngay
                material_subgroup = str(row.get('material_subgroup', 'Ch∆∞a ph√¢n lo·∫°i')).strip() if pd.notna(row.get('material_subgroup')) else 'Ch∆∞a ph√¢n lo·∫°i'
                
                material_subprice = row.get('material_subprice')
                if pd.notna(material_subprice) and isinstance(material_subprice, str):
                    try:
                        json.loads(material_subprice)
                        material_subprice_json = material_subprice
                    except:
                        material_subprice_json = None
                else:
                    material_subprice_json = None
                
                unit = str(row.get('unit', '')).strip() if pd.notna(row.get('unit')) else None
                image_url = str(row.get('image_url', '')).strip() if pd.notna(row.get('image_url')) else None
                
                sql = """
                    INSERT INTO materials (
                        id_sap, material_name, material_group, material_subgroup,
                        material_subprice, unit, image_url
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (id_sap) DO UPDATE SET 
                        material_name = EXCLUDED.material_name,
                        material_group = EXCLUDED.material_group,
                        material_subgroup = EXCLUDED.material_subgroup,
                        material_subprice = EXCLUDED.material_subprice,
                        unit = EXCLUDED.unit,
                        image_url = EXCLUDED.image_url,
                        updated_at = NOW()
                """
                
                cur.execute(sql, (
                    id_sap, material_name, material_group, material_subgroup,
                    material_subprice_json, unit, image_url
                ))
                
                imported += 1
                
            except Exception as e:
                errors.append(f"Row {idx+2}: {str(e)[:100]}")
        
        conn.commit()
        conn.close()
        
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM materials 
            WHERE material_subgroup = 'Ch∆∞a ph√¢n lo·∫°i'
        """)
        pending_count = cur.fetchone()[0]
        conn.close()
        
        message = f"‚úÖ Import th√†nh c√¥ng {imported}/{len(df)} materials"
        if pending_count > 0:
            message += f"\n\n‚è≥ C√≥ {pending_count} v·∫≠t li·ªáu ch∆∞a ph√¢n lo·∫°i."
            message += f"\nüí° D√πng n√∫t 'ü§ñ Auto Classify Materials' ƒë·ªÉ ph√¢n lo·∫°i."
        
        return {
            "message": message,
            "imported": imported,
            "total": len(df),
            "pending_classification": pending_count,
            "errors": errors[:10] if errors else []
        }
        
    except Exception as e:
        return {"message": f"‚ùå L·ªói: {str(e)}"}



@app.post("/import/product-materials")
async def import_product_materials(file: UploadFile = File(...)):
    """
    [V4.5] Import ƒë·ªãnh m·ª©c - T·ª± ƒë·ªông t·∫°o v·∫≠t li·ªáu thi·∫øu (Placeholder)
    - N·∫øu m√£ v·∫≠t li·ªáu ch∆∞a c√≥ trong kho -> T·ª± ƒë·ªông t·∫°o m·ªõi ƒë·ªÉ tr√°nh l·ªói
    - Fix l·ªói ƒëu√¥i .0
    """
    try:
        contents = await file.read()
        df = pd.read_csv(io.BytesIO(contents))
        
        # Chu·∫©n h√≥a t√™n c·ªôt
        df.columns = df.columns.str.strip().str.lower()
        
        required = ['product_headcode']
        missing = [col for col in required if col not in df.columns]
        
        if missing:
            return {
                "message": f"‚ùå Thi·∫øu c·ªôt b·∫Øt bu·ªôc: {', '.join(missing)}",
                "required_columns": required,
                "your_columns": list(df.columns)
            }
        
        conn = get_db()
        cur = conn.cursor()
        
        imported = 0
        skipped = 0
        auto_created_materials = 0 # ƒê·∫øm s·ªë v·∫≠t li·ªáu ƒë∆∞·ª£c t·∫°o t·ª± ƒë·ªông
        errors = []
        
        # Pre-load d·ªØ li·ªáu ƒë·ªÉ check nhanh
        cur.execute("SELECT headcode FROM products")
        existing_products = {row[0] for row in cur.fetchall()}
        
        cur.execute("SELECT id_sap FROM materials")
        existing_materials = {row[0] for row in cur.fetchall()}

        # H√†m l√†m s·∫°ch ID
        def clean_id(val):
            if pd.isna(val) or val == '':
                return ""
            s = str(val).strip()
            if s.endswith('.0'):
                return s[:-2]
            return s
        
        for idx, row in df.iterrows():
            savepoint_name = f"sp_{idx}"
            cur.execute(f"SAVEPOINT {savepoint_name}")
            
            try:
                # 1. X·ª≠ l√Ω Product (V·∫´n b·∫Øt bu·ªôc ph·∫£i c√≥ tr∆∞·ªõc)
                product_headcode = clean_id(row.get('product_headcode'))
                
                if not product_headcode or product_headcode.lower() == 'nan':
                    errors.append(f"Row {idx+2}: Thi·∫øu Product Headcode")
                    continue 

                if product_headcode not in existing_products:
                    # T√πy ch·ªçn: C√≥ th·ªÉ mu·ªën t·ª± t·∫°o Product lu√¥n, nh∆∞ng th∆∞·ªùng Product c·∫ßn ki·ªÉm so√°t ch·∫∑t h∆°n
                    raise ValueError(f"Product '{product_headcode}' ch∆∞a c√≥ trong h·ªá th·ªëng")

                # 2. X·ª≠ l√Ω Material (T·ª± ƒë·ªông t·∫°o n·∫øu thi·∫øu)
                material_id_sap = clean_id(row.get('material_id_sap'))
                
                if not material_id_sap or material_id_sap.lower() == 'nan':
                    skipped += 1
                    cur.execute(f"RELEASE SAVEPOINT {savepoint_name}")
                    continue 

                # --- LOGIC M·ªöI: T·ª∞ ƒê·ªòNG T·∫†O V·∫¨T LI·ªÜU N·∫æU THI·∫æU ---
                if material_id_sap not in existing_materials:
                    # T·∫°o v·∫≠t li·ªáu t·∫°m
                    temp_name = f"V·∫≠t li·ªáu m·ªõi {material_id_sap}"
                    
                    cur.execute("""
                        INSERT INTO materials (id_sap, material_name, material_group, material_subgroup)
                        VALUES (%s, %s, %s, %s)
                        ON CONFLICT (id_sap) DO NOTHING
                    """, (material_id_sap, temp_name, "Auto-Created", "Ch·ªù c·∫≠p nh·∫≠t"))
                    
                    # C·∫≠p nh·∫≠t v√†o set ƒë·ªÉ c√°c d√≤ng sau kh√¥ng insert l·∫°i
                    existing_materials.add(material_id_sap)
                    auto_created_materials += 1
                # --------------------------------------------------

                # 3. Insert v√†o b·∫£ng ƒë·ªãnh m·ª©c
                quantity = float(row['quantity']) if pd.notna(row.get('quantity')) else 0
                unit = str(row.get('unit', '')).strip() if pd.notna(row.get('unit')) else None
                
                sql = """
                    INSERT INTO product_materials (product_headcode, material_id_sap, quantity, unit)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (product_headcode, material_id_sap) DO UPDATE SET
                        quantity = EXCLUDED.quantity,
                        unit = EXCLUDED.unit,
                        updated_at = NOW()
                """
                
                cur.execute(sql, (product_headcode, material_id_sap, quantity, unit))
                
                cur.execute(f"RELEASE SAVEPOINT {savepoint_name}")
                imported += 1
                
            except Exception as e:
                cur.execute(f"ROLLBACK TO SAVEPOINT {savepoint_name}")
                errors.append(f"Row {idx+2}: {str(e)}")

        conn.commit()
        conn.close()
        
        msg = f"‚úÖ Import th√†nh c√¥ng {imported} d√≤ng."
        if auto_created_materials > 0:
            msg += f"\nüÜï ƒê√£ t·ª± ƒë·ªông t·∫°o m·ªõi {auto_created_materials} m√£ v·∫≠t li·ªáu (ch∆∞a c√≥ th√¥ng tin)."
        if skipped > 0:
            msg += f"\n‚ö†Ô∏è B·ªè qua {skipped} d√≤ng do kh√¥ng c√≥ m√£ v·∫≠t li·ªáu."
            
        return {
            "message": msg,
            "imported": imported,
            "auto_created_materials": auto_created_materials,
            "skipped": skipped,
            "total_rows": len(df),
            "errors": errors[:10] if errors else []
        }
        
    except Exception as e:
        return {"message": f"‚ùå L·ªói h·ªá th·ªëng: {str(e)}"}


# ========================================
# [3] NEW BATCH CLASSIFICATION ENDPOINTS
# Th√™m 2 endpoints m·ªõi ƒë·ªÉ classify sau khi import
# ========================================

@app.post("/classify-products")
def classify_pending_products():
    """
    ü§ñ Ph√¢n lo·∫°i H√ÄNG LO·∫†T c√°c s·∫£n ph·∫©m ch∆∞a ph√¢n lo·∫°i
    Batch size: 8 s·∫£n ph·∫©m/l·∫ßn (tr√°nh qu√° d√†i response)
    """
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # L·∫•y s·∫£n ph·∫©m ch∆∞a ph√¢n lo·∫°i
        cur.execute("""
            SELECT headcode, id_sap, product_name 
            FROM products 
            WHERE category = 'Ch∆∞a ph√¢n lo·∫°i' 
               OR sub_category = 'Ch∆∞a ph√¢n lo·∫°i'
               OR material_primary = 'Ch∆∞a x√°c ƒë·ªãnh'
            LIMIT 100
        """)
        
        pending_products = cur.fetchall()
        
        if not pending_products:
            conn.close()
            return {
                "message": "‚úÖ T·∫•t c·∫£ s·∫£n ph·∫©m ƒë√£ ƒë∆∞·ª£c ph√¢n lo·∫°i!",
                "classified": 0,
                "total": 0,
                "remaining": 0
            }
        
        total_pending = len(pending_products)
        classified = 0
        errors = []
        
        BATCH_SIZE = 8  # Gemini x·ª≠ l√Ω t·ªët v·ªõi 5-10 items
        
        for i in range(0, len(pending_products), BATCH_SIZE):
            batch = pending_products[i:i+BATCH_SIZE]
            
            # Chu·∫©n b·ªã input cho batch classification
            batch_input = [{
                'id_sap': p['id_sap'],
                'name': p['product_name']
            } for p in batch]
            
            print(f"ü§ñ Classifying batch {i//BATCH_SIZE + 1} ({len(batch)} products)...")
            
            try:
                # G·ªåI BATCH CLASSIFICATION
                results = batch_classify_products(batch_input)
                
                # C·∫≠p nh·∫≠t v√†o DB
                for j, result in enumerate(results):
                    try:
                        cur.execute("""
                            UPDATE products 
                            SET category = %s,
                                sub_category = %s,
                                material_primary = %s,
                                updated_at = NOW()
                            WHERE headcode = %s
                        """, (
                            result['category'],
                            result['sub_category'],
                            result['material_primary'],
                            batch[j]['headcode']
                        ))
                        classified += 1
                    except Exception as e:
                        errors.append(f"{batch[j]['headcode']}: {str(e)[:50]}")
                
                conn.commit()
                
                # Delay gi·ªØa c√°c batch ƒë·ªÉ tr√°nh rate limit
                if i + BATCH_SIZE < len(pending_products):
                    time.sleep(4)
                
            except Exception as e:
                print(f"‚ùå Batch {i//BATCH_SIZE + 1} failed: {e}")
                errors.append(f"Batch {i//BATCH_SIZE + 1}: {str(e)[:100]}")
                # Ti·∫øp t·ª•c v·ªõi batch ti·∫øp theo
                continue
        
        conn.close()
        
        # Ki·ªÉm tra c√≤n bao nhi√™u ch∆∞a ph√¢n lo·∫°i
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM products 
            WHERE category = 'Ch∆∞a ph√¢n lo·∫°i' 
            OR sub_category = 'Ch∆∞a ph√¢n lo·∫°i'
            OR material_primary = 'Ch∆∞a x√°c ƒë·ªãnh'
        """)
        remaining = cur.fetchone()[0]
        conn.close()
        
        return {
            "message": f"‚úÖ ƒê√£ ph√¢n lo·∫°i {classified}/{total_pending} s·∫£n ph·∫©m",
            "classified": classified,
            "total": total_pending,
            "remaining": remaining,
            "errors": errors[:10] if errors else []
        }
        
    except Exception as e:
        return {
            "message": f"‚ùå L·ªói: {str(e)}",
            "classified": 0,
            "total": 0,
            "remaining": 0
        }


@app.post("/classify-materials")
def classify_pending_materials():
    """
    ü§ñ Ph√¢n lo·∫°i H√ÄNG LO·∫†T c√°c v·∫≠t li·ªáu ch∆∞a ph√¢n lo·∫°i
    """
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT id_sap, material_name, material_group
            FROM materials 
            WHERE material_subgroup = 'Ch∆∞a ph√¢n lo·∫°i'
            LIMIT 100
        """)
        
        pending_materials = cur.fetchall()
        
        if not pending_materials:
            conn.close()
            return {
                "message": "‚úÖ T·∫•t c·∫£ v·∫≠t li·ªáu ƒë√£ ƒë∆∞·ª£c ph√¢n lo·∫°i!",
                "classified": 0,
                "total": 0,
                "remaining": 0
            }
        
        total_pending = len(pending_materials)
        classified = 0
        errors = []
        
        BATCH_SIZE = 10
        
        for i in range(0, len(pending_materials), BATCH_SIZE):
            batch = pending_materials[i:i+BATCH_SIZE]
            
            batch_input = [{
                'id_sap': m['id_sap'],
                'name': m['material_name']
            } for m in batch]
            
            print(f"ü§ñ Classifying materials batch {i//BATCH_SIZE + 1} ({len(batch)} items)...")
            
            try:
                results = batch_classify_materials(batch_input)
                
                for j, result in enumerate(results):
                    try:
                        cur.execute("""
                            UPDATE materials 
                            SET material_subgroup = %s,
                                updated_at = NOW()
                            WHERE id_sap = %s
                        """, (
                            result['material_subgroup'],
                            batch[j]['id_sap']
                        ))
                        classified += 1
                    except Exception as e:
                        errors.append(f"{batch[j]['id_sap']}: {str(e)[:50]}")
                
                conn.commit()
                
                if i + BATCH_SIZE < len(pending_materials):
                    time.sleep(4)
                
            except Exception as e:
                print(f"‚ùå Materials batch {i//BATCH_SIZE + 1} failed: {e}")
                errors.append(f"Batch {i//BATCH_SIZE + 1}: {str(e)[:100]}")
                continue
        
        conn.close()
        
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM materials 
            WHERE material_subgroup = 'Ch∆∞a ph√¢n lo·∫°i'
        """)
        remaining = cur.fetchone()[0]
        conn.close()
        
        return {
            "message": f"‚úÖ ƒê√£ ph√¢n lo·∫°i {classified}/{total_pending} v·∫≠t li·ªáu",
            "classified": classified,
            "total": total_pending,
            "remaining": remaining,
            "errors": errors[:10] if errors else []
        }
        
    except Exception as e:
        return {
            "message": f"‚ùå L·ªói: {str(e)}",
            "classified": 0,
            "total": 0,
            "remaining": 0
        }





# ========================================
# GENERATE EMBEDDINGS
# ========================================

@app.post("/generate-embeddings")
def generate_product_embeddings():
    """T√°¬∫¬°o embeddings cho products"""
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    cur.execute("""
        SELECT headcode, product_name, category, sub_category, material_primary
        FROM products 
        WHERE name_embedding IS NULL OR description_embedding IS NULL
        LIMIT 100
    """)
    
    products = cur.fetchall()
    
    if not products:
        conn.close()
        return {"message": "√¢≈ì‚Ä¶ T√°¬∫¬•t c√°¬∫¬£ products √Ñ‚Äò√É¬£ c√É¬≥ embeddings"}
    
    success = 0
    errors = []
    
    for prod in products:
        try:
            name_text = f"{prod['product_name']}"
            name_emb = generate_embedding(name_text)
            
            desc_text = f"{prod['product_name']} {prod.get('category', '')} {prod.get('sub_category', '')} {prod.get('material_primary', '')}"
            desc_emb = generate_embedding(desc_text)
            
            if name_emb and desc_emb:
                cur.execute("""
                    UPDATE products 
                    SET name_embedding = %s, description_embedding = %s, updated_at = NOW()
                    WHERE headcode = %s
                """, (name_emb, desc_emb, prod['headcode']))
                
                success += 1
                time.sleep(0.5)
            
        except Exception as e:
            errors.append(f"{prod['headcode']}: {str(e)[:50]}")
    
    conn.commit()
    conn.close()
    
    return {
        "message": f"√¢≈ì‚Ä¶ √Ñ √É¬£ t√°¬∫¬°o embeddings cho {success}/{len(products)} products",
        "success": success,
        "total": len(products),
        "errors": errors[:5] if errors else []
    }

@app.post("/generate-material-embeddings")
def generate_material_embeddings():
    """T√°¬∫¬°o embeddings cho materials"""
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    cur.execute("""
        SELECT id_sap, material_name, material_group, material_subgroup
        FROM materials 
        WHERE name_embedding IS NULL OR description_embedding IS NULL
        LIMIT 100
    """)
    
    materials = cur.fetchall()
    
    if not materials:
        conn.close()
        return {"message": "√¢≈ì‚Ä¶ T√°¬∫¬•t c√°¬∫¬£ materials √Ñ‚Äò√É¬£ c√É¬≥ embeddings"}
    
    success = 0
    errors = []
    
    for mat in materials:
        try:
            name_text = f"{mat['material_name']}"
            name_emb = generate_embedding(name_text)
            
            desc_text = f"{mat['material_name']} {mat.get('material_group', '')} {mat.get('material_subgroup', '')}"
            desc_emb = generate_embedding(desc_text)
            
            if name_emb and desc_emb:
                cur.execute("""
                    UPDATE materials 
                    SET name_embedding = %s, description_embedding = %s, updated_at = NOW()
                    WHERE id_sap = %s
                """, (name_emb, desc_emb, mat['id_sap']))
                
                success += 1
                time.sleep(0.5)
            
        except Exception as e:
            errors.append(f"{mat['id_sap']}: {str(e)[:50]}")
    
    conn.commit()
    conn.close()
    
    return {
        "message": f"√¢≈ì‚Ä¶ √Ñ √É¬£ t√°¬∫¬°o embeddings cho {success}/{len(materials)} materials",
        "success": success,
        "total": len(materials),
        "errors": errors[:5] if errors else []
    }

# ========================================
# DEBUG ENDPOINTS
# ========================================

@app.get("/debug/products")
def debug_products():
    """Debug info v√°¬ª  products"""
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    cur.execute("SELECT COUNT(*) as total FROM products")
    total = cur.fetchone()['total']
    
    cur.execute("SELECT COUNT(*) as with_emb FROM products WHERE description_embedding IS NOT NULL")
    with_emb = cur.fetchone()['with_emb']
    
    cur.execute("SELECT category, COUNT(*) as count FROM products GROUP BY category ORDER BY count DESC")
    by_category = cur.fetchall()
    
    conn.close()
    
    return {
        "total_products": total,
        "with_embeddings": with_emb,
        "coverage_percent": round(with_emb / total * 100, 1) if total > 0 else 0,
        "by_category": [dict(c) for c in by_category]
    }

@app.get("/debug/materials")
def debug_materials():
    """Debug info v√°¬ª  materials"""
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    cur.execute("SELECT COUNT(*) as total FROM materials")
    total = cur.fetchone()['total']
    
    cur.execute("SELECT COUNT(*) as with_emb FROM materials WHERE description_embedding IS NOT NULL")
    with_emb = cur.fetchone()['with_emb']
    
    cur.execute("SELECT material_group, COUNT(*) as count FROM materials GROUP BY material_group ORDER BY count DESC")
    by_group = cur.fetchall()
    
    conn.close()
    
    return {
        "total_materials": total,
        "with_embeddings": with_emb,
        "coverage_percent": round(with_emb / total * 100, 1) if total > 0 else 0,
        "by_group": [dict(g) for g in by_group]
    }

@app.get("/debug/chat-history")
def debug_chat_history():
    """Xem l√°¬ª‚Äπch s√°¬ª¬≠ chat g√°¬∫¬ßn √Ñ‚Äò√É¬¢y"""
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    cur.execute("""
        SELECT 
            session_id,
            user_message,
            intent,
            result_count,
            created_at
        FROM chat_history
        ORDER BY created_at DESC
        LIMIT 20
    """)
    
    history = cur.fetchall()
    conn.close()
    
    return {
        "recent_chats": [dict(h) for h in history]
    }

# ========================================
# [4] UPDATE ROOT ENDPOINT
# C·∫≠p nh·∫≠t danh s√°ch endpoints
# ========================================

@app.get("/")
def root():
    return {
        "app": "AA Corporation Chatbot API", 
        "version": "4.1",
        "status": "Running",
        "features": [
            "‚úÖ Queue-based batch classification",
            "‚úÖ Import tr∆∞·ªõc, classify sau",
            "‚úÖ Batch size 8-10 items/call",
            "‚úÖ Ti·∫øt ki·ªám quota Gemini",
            "‚úÖ NULL safety 100%"
        ],
        "endpoints": {
            "chat": "POST /chat",
            "search_image": "POST /search-image",
            "import_products": "POST /import/products",
            "import_materials": "POST /import/materials",
            "import_pm": "POST /import/product-materials",
            "classify_products": "POST /classify-products üÜï",
            "classify_materials": "POST /classify-materials üÜï",
            "generate_embeddings": "POST /generate-embeddings",
            "generate_material_embeddings": "POST /generate-material-embeddings",
            "debug": "GET /debug/products, /debug/materials, /debug/chat-history"
        }
    }

@app.get("/history/{session_id}")
def get_session_history(session_id: str):
    """Xem l·ªãch s·ª≠ c·ªßa 1 user - V4.6"""
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        sql = """
            SELECT 
                user_message,
                intent,
                search_type,
                expanded_query,
                extracted_keywords,
                result_count,
                created_at
            FROM chat_history
            WHERE session_id = %s
            ORDER BY created_at DESC
            LIMIT 20
        """
        
        cur.execute(sql, (session_id,))
        history = cur.fetchall()
        conn.close()
        
        return {
            "session_id": session_id,
            "total_queries": len(history),
            "history": [dict(h) for h in history]
        }
    except Exception as e:
        return {"error": str(e)}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)