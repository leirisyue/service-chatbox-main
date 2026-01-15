import json
import time
from io import BytesIO
from typing import Dict, List

import google.generativeai as genai
import numpy as np
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from psycopg2.extras import RealDictCursor

from config import settings

from .embeddingapi import generate_embedding_qwen


def get_db():
    return psycopg2.connect(**settings.DB_CONFIG)

def format_suggested_prompts(prompts: list[str]) -> str:
    if not prompts:
        return ""
    return "\n".join([f"• {p}" for p in prompts])

def extract_product_keywords(query: str) -> list:
    """Trích xuất từ khóa quan trọng, bao gồm cụm từ"""
    materials = ["gỗ teak", "gỗ sồi", "gỗ walnut", "đá marble", "đá granite", 
                    "da thật", "da bò", "vải linen", "kim loại", "teak", "oak", 
                    "walnut", "marble", "granite", "leather"]
    
    contexts = ["nhà bếp", "phòng khách", "phòng ngủ", "văn phòng",
                "kitchen", "living room", "dining", "coffee", "bar",
                "bàn ăn", "bàn trà", "bàn làm việc"]
    
    shapes = ["tròn", "vuông", "chữ nhật", "oval", "l-shape", 
                "round", "square", "rectangular"]
    
    table_types = ["bàn làm việc", "bàn ăn", "bàn trà", "bàn coffee", 
                    "bàn học", "bàn máy tính", "working table", "desk", 
                    "dining table", "coffee table", "study table"]
    
    chair_types = ["ghế ăn", "ghế bar", "ghế sofa", "ghế văn phòng",
                    "dining chair", "bar chair", "office chair"]
    
    types = ["bàn", "ghế", "tủ", "giường", "sofa", "kệ", "đèn",
                "table", "chair", "cabinet", "bed", "shelf", "lamp"]
    
    query_lower = query.lower()
    keywords = []
    
    for word_list in [table_types, chair_types, materials, contexts, shapes]:
        for word in word_list:
            if word in query_lower:
                keywords.append(word)
    
    for word in types:
        if word in query_lower:
            if not any(word in kw for kw in keywords):
                keywords.append(word)
    
    keywords = list(set(keywords))
    if keywords:
        print(f"INFO: Keywords => {keywords}")
    return keywords

def auto_classify_product(product_name: str, id_sap: str = "") -> Dict:
    """Tự động phân loại sản phẩm bằng AI"""
    model = genai.GenerativeModel("gemini-2.5-flash")
    
    prompt = f"""
                Bạn là chuyên gia phân loại sản phẩm nội thất cao cấp.
                INPUT:
                - Tên sản phẩm: "{product_name}"
                - Mã SAP: "{id_sap}"
                NHIỆM VỤ: Phân tích và phân loại sản phẩm theo 3 tiêu chí:
                1. **category** (Danh mục chính):
                - Bàn (Table)
                - Ghế (Chair) 
                - Sofa
                - Tủ (Cabinet)
                - Giường (Bed)
                - Đèn (Lamp)
                - Kệ (Shelf)
                - Bàn làm việc (Desk)
                - Khác (Other)
                2. **sub_category** (Danh mục phụ - cụ thể hơn):
                VD: "Bàn ăn", "Bàn coffee", "Ghế bar", "Ghế ăn", "Sofa góc", "Tủ quần áo", "Đèn bàn", "Đèn trần"...
                3. **material_primary** (Vật liệu chính):
                - Gỗ (Wood)
                - Da (Leather)
                - Vải (Fabric)
                - Kim loại (Metal)
                - Đá (Stone)
                - Kính (Glass)
                - Nhựa (Plastic)
                - Mây tre (Rattan)
                - Hỗn hợp (Mixed)
                OUTPUT JSON ONLY (no markdown, no backticks):
                {{
                "category": "...",
                "sub_category": "...",
                "material_primary": "..."
                }}
        """
    
    response_text = call_gemini_with_retry(model, prompt)
    
    if not response_text:
        return {
            "category": "Chưa phân loại",
            "sub_category": "Chưa phân loại", 
            "material_primary": "Chưa xác định"
        }
    
    try:
        clean = response_text.strip()
        if "```json" in clean:
            clean = clean.split("```json")[1].split("```")[0].strip()
        elif "```" in clean:
            clean = clean.split("```")[1].split("```")[0].strip()
        
        result = json.loads(clean)
        return result
    except:
        return {
            "category": "Chưa phân loại",
            "sub_category": "Chưa phân loại",
            "material_primary": "Chưa xác định"
        }

def auto_classify_material(material_name: str, id_sap: str = "") -> Dict:
    """Tự động phân loại vật liệu bằng AI"""
    model = genai.GenerativeModel("gemini-2.5-flash")
    
    prompt = f"""
                Phân loại nguyên vật liệu nội thất:
                Tên: "{material_name}"
                Mã: "{id_sap}"
                Xác định:
                1. **material_group**: Gỗ, Da, Vải, Đá, Kim loại, Kính, Nhựa, Sơn, Keo, Phụ kiện, Khác
                2. **material_subgroup**: Nhóm con cụ thể (VD: "Gỗ tự nhiên", "Da thật", "Vải cao cấp"...)
                OUTPUT JSON ONLY:
                {{
                    "material_group": "...",
                    "material_subgroup": "..."
                }}
        """
    
    response_text = call_gemini_with_retry(model, prompt)
    
    if not response_text:
        return {
            "material_group": "Chưa phân loại",
            "material_subgroup": "Chưa phân loại"
        }
    
    try:
        clean = response_text.strip()
        if "```json" in clean:
            clean = clean.split("```json")[1].split("```")[0].strip()
        elif "```" in clean:
            clean = clean.split("```")[1].split("```")[0].strip()
        
        result = json.loads(clean)
        return result
    except:
        return {
            "material_group": "Chưa phân loại",
            "material_subgroup": "Chưa phân loại"
        }

def search_materials_for_product(product_query: str, params: Dict):
    """
    🔍 TÌM VẬT LIỆU ĐỂ LÀM SẢN PHẨM CỤ THỂ
    Ví dụ: "Vật liệu làm bàn tròn", "Nguyên liệu ghế sofa"
    
    Logic:
    1. Tìm products phù hợp với query
    2. JOIN product_materials để lấy materials được dùng
    3. Aggregate + rank theo tần suất
    """
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    print(f"INFO: Cross-table search: Materials for '{product_query}'")
    
    # Bước 1: Tìm products phù hợp
    product_vector = generate_embedding_qwen(product_query)
    
    if not product_vector:
        conn.close()
        return {"materials": [], "search_method": "failed"}
    
    try:
        cur.execute("""
            SELECT 
                headcode,
                product_name,
                category,
                (description_embedding <=> %s::vector) as distance
            FROM products_qwen
            WHERE description_embedding IS NOT NULL
            ORDER BY distance ASC
            LIMIT 10
        """, [product_vector])
        
        matched_products = cur.fetchall()
        
        if not matched_products:
            conn.close()
            return {"materials": [], "search_method": "no_products_found"}
        
        product_headcodes = [p['headcode'] for p in matched_products]
        product_names = [p['product_name'] for p in matched_products]
        
        print(f"SUCCESS: Found {len(product_headcodes)} matching products: {product_names[:3]}")
        
        # Bước 2: Lấy materials được dùng trong products này
        material_filter = ""
        filter_params = []
        
        if params.get("material_group"):
            material_filter = "AND m.material_group ILIKE %s"
            filter_params.append(f"%{params['material_group']}%")
        
        sql = f"""
            SELECT 
                m.id_sap,
                m.material_name,
                m.material_group,
                m.material_subgroup,
                m.material_subprice,
                m.unit,
                m.image_url,
                COUNT(DISTINCT pm.product_headcode) as usage_count,
                SUM(pm.quantity) as total_quantity,
                array_agg(DISTINCT p.product_name) as used_in_products
            FROM {settings.MATERIALS_TABLE} m
            INNER JOIN product_materials pm ON m.id_sap = pm.material_id_sap
            INNER JOIN products_qwen p ON pm.product_headcode = p.headcode
            WHERE p.headcode = ANY(%s)
            {material_filter}
            GROUP BY m.id_sap, m.material_name, m.material_group, 
                    m.material_subgroup, m.material_subprice, m.unit, m.image_url
            ORDER BY usage_count DESC, m.material_name ASC
            LIMIT 15
        """
        
        cur.execute(sql, [product_headcodes] + filter_params)
        results = cur.fetchall()
        
        conn.close()
        
        if not results:
            return {
                "materials": [],
                "search_method": "cross_table_no_materials",
                "matched_products": product_names
            }
        
        materials_with_context = []
        for mat in results:
            mat_dict = dict(mat)
            mat_dict['price'] = get_latest_material_price(mat['material_subprice'])
            mat_dict['used_in_products_list'] = mat['used_in_products'][:5]  # Top 5
            materials_with_context.append(mat_dict)
        
        print(f"SUCCESS: Found {len(materials_with_context)} materials used in these products")
        
        return {
            "materials": materials_with_context,
            "search_method": "cross_table_product_to_material",
            "matched_products": product_names[:5],
            "explanation": f"Vật liệu thường dùng cho: {', '.join(product_names[:3])}"
        }
        
    except Exception as e:
        print(f"ERROR: Cross-table materials search failed: {e}")
        conn.close()
        return {"materials": [], "search_method": "cross_table_error"}

def get_adaptive_threshold(query: str) -> float:
    """
    Tự động điều chỉnh threshold:
    - Query dài, cụ thể → threshold thấp (0.75)
    - Query ngắn, chung chung → threshold cao (0.90)
    """
    words = query.split()
    
    if len(words) >= 8:
        return 0.75  # Query dài → dễ dãi hơn
    elif len(words) >= 5:
        return 0.82
    else:
        return 0.90  

def format_search_results(results):
    """Format results thành cấu trúc chuẩn"""
    products = []
    for row in results:
        products.append({
            "headcode": row["headcode"],
            "product_name": row["product_name"],
            "category": row.get("category"),
            "sub_category": row.get("sub_category"),
            "material_primary": row.get("material_primary"),
            "project": row.get("project"),
            "project_id": row.get("project_id"),
            "similarity": round(1 - row["distance"], 3) if "distance" in row else None
        })
    return products

def call_gemini_with_retry(model, prompt, max_retries=3, timeout=20):
    """Gọi Gemini với retry logic và timeout"""
    import signal
    
    def timeout_handler(signum, frame):
        raise TimeoutError("Gemini API timeout")
    
    for attempt in range(max_retries):
        try:
            # Set timeout for this attempt (only on Unix systems)
            if hasattr(signal, 'SIGALRM'):
                signal.signal(signal.SIGALRM, timeout_handler)
                signal.alarm(timeout)
            
            response = model.generate_content(prompt, request_options={"timeout": timeout})
            
            # Cancel alarm
            if hasattr(signal, 'SIGALRM'):
                signal.alarm(0)
            
            if response.text:
                return response.text
        except TimeoutError:
            print(f"WARNING: Gemini timeout after {timeout}s on attempt {attempt + 1}")
            if attempt == max_retries - 1:
                return None
            continue
        except Exception as e:
            # Cancel alarm on error
            if hasattr(signal, 'SIGALRM'):
                signal.alarm(0)
            
            if "429" in str(e) or "quota" in str(e).lower():
                wait_time = 5 * (2 ** attempt)
                print(f"INFO: Quota exceeded. Đợi {wait_time}s...")
                time.sleep(wait_time)
                continue
            print(f"ERROR Gemini: {e}")
            return None
    return None

def calculate_product_total_cost(headcode: str) -> float:
    """Tính tổng chi phí (total_cost) cho một sản phẩm"""
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    sql = """
        SELECT 
            m.material_subprice,
            pm.quantity
        FROM product_materials pm
        INNER JOIN materials m ON pm.material_id_sap = m.id_sap
        WHERE pm.product_headcode = %s
    """
    try:
        cur.execute(sql, (headcode,))
        materials = cur.fetchall()
    except Exception as e:
        print(f"ERROR: Query error in calculate_product_total_cost for {headcode}: {e}")
        conn.close()
        return 0.0
    conn.close()
    if not materials:
        return 0.0
    
    material_cost = 0
    for mat in materials:
        quantity = float(mat['quantity']) if mat['quantity'] else 0.0
        latest_price = get_latest_material_price(mat['material_subprice'])
        material_cost += quantity * latest_price  # Sửa lỗi: cộng dồn material_cost

    labor_cost = material_cost * 0.20
    overhead_cost = material_cost * 0.15
    profit_margin = material_cost * 0.25
    
    total_cost = material_cost + labor_cost + overhead_cost + profit_margin
    return total_cost

def search_products_hybrid(params: Dict):
    """HYBRID: Vector + Keyword với từ CHÍNH bắt buộc khớp, từ PHỤ tìm gần giống"""
    import signal
    
    def timeout_handler(signum, frame):
        raise TimeoutError("Search timeout")
    
    # Set timeout cho toàn bộ search operation (20 giây)
    if hasattr(signal, 'SIGALRM'):
        signal.signal(signal.SIGALRM, timeout_handler)
        signal.alarm(20)
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    # 1. Chuẩn bị query
    if params.get("keywords_vector"):
        base = params["keywords_vector"]
    else:
        parts = [params.get("category", ""), params.get("sub_category", ""), 
                params.get("material_primary", "")]
        base = " ".join([p for p in parts if p]) or "nội thất"
    
    # ✅ XỬ LÝ ĐẶC BIỆT: Tìm "danh sách sản phẩm" - lấy 1 sản phẩm mỗi loại
    query_lower = base.lower()
    if "danh sách" in query_lower or "product list" in query_lower or "product catalog" in query_lower:
        print(f"🔍 Special query detected: Product list - returning one product per category")
        try:
            sql = """
                SELECT DISTINCT ON (category) 
                    headcode, product_name, category, sub_category, 
                    material_primary, project, project_id
                FROM products_qwen
                WHERE category IS NOT NULL
                    AND category != ''
                ORDER BY category, headcode
                LIMIT 10
            """
            cur.execute(sql)
            products = cur.fetchall()
            
            if products:
                result = []
                for p in products:
                    result.append({
                        "headcode": p["headcode"],
                        "product_name": p["product_name"],
                        "category": p.get("category"),
                        "sub_category": p.get("sub_category"),
                        "material_primary": p.get("material_primary"),
                        "project": p.get("project"),
                        "project_id": p.get("project_id"),
                        "similarity": 0.9,
                        "final_score": 0.9
                    })
                
                conn.close()
                if hasattr(signal, 'SIGALRM'):
                    signal.alarm(0)
                
                print(f"SUCCESS: Found {len(result)} products (one per category)")
                return {
                    "products": result,
                    "search_method": "product_list_by_category",
                    "expanded_query": base
                }
        except Exception as e:
            print(f"ERROR: Error in product list query: {e}")
            # Fall through to normal search if error
    
    # print(f"\n🔍 Query: {base}")
    
    # 2. AI Expansion với timeout ngắn hơn
    expanded = expand_search_query(base, params)
    
    # 3. Extract keywords
    keywords = extract_product_keywords(expanded)
    
    # 4. Tách từ trong query gốc
    original_words = [w.strip().lower() for w in base.split() if len(w.strip()) > 1]
    
    # 5. XÁC ĐỊNH TỪ CHÍNH (loại sản phẩm) - PHẢI KHỚP CHÍNH XÁC
    main_product_types = ["bàn", "ghế", "tủ", "giường", "sofa", "kệ", "đèn", "gương",
                          "table", "chair", "cabinet", "bed", "shelf", "lamp", "mirror"]
    
    main_word = None
    secondary_words = []
    
    for word in original_words:
        if word in main_product_types:
            main_word = word
            break
    
    # Nếu không tìm thấy từ chính trong danh sách, lấy từ đầu tiên làm từ chính
    if not main_word and original_words:
        main_word = original_words[0]
    
    # Các từ còn lại là từ phụ
    secondary_words = [w for w in original_words if w != main_word]
    
    print(f"🔍 Main word (REQUIRED): '{main_word}' | Secondary: {secondary_words}")
    
    # 6. Vector
    vector = generate_embedding_qwen(expanded)
    if not vector:
        conn.close()
        if hasattr(signal, 'SIGALRM'):
            signal.alarm(0)
        return {"products": [], "search_method": "failed", "error": "no_vector"}
    
    # 7. BƯỚC 1: Tìm trong DATABASE với TỪ CHÍNH (keyword search)
    try:
        if not main_word:
            print("⚠️ No main word detected, returning empty")
            conn.close()
            if hasattr(signal, 'SIGALRM'):
                signal.alarm(0)
            return {"products": [], "search_method": "no_main_word"}
        
        # BƯỚC 1: Query database với từ CHÍNH - CHỈ TÌM TRONG PRODUCT_NAME
        print(f"STEP 1: Query DB with main word: '{main_word}'")
        
        sql_step1 = """
            SELECT headcode, product_name, category, sub_category, 
                   material_primary, project, project_id, description_embedding
            FROM products_qwen
            WHERE description_embedding IS NOT NULL
                AND product_name ILIKE %s
            LIMIT 100
        """
        
        cur.execute(sql_step1, [f"%{main_word}%"])
        candidates = cur.fetchall()
        
        if not candidates:
            print(f"ERROR: No products found with main word '{main_word}' in product_name")
            conn.close()
            if hasattr(signal, 'SIGALRM'):
                signal.alarm(0)
            return {
                "products": [],
                "search_method": "no_candidates_with_main_word",
                "main_word": main_word
            }
    
    except TimeoutError:
        print(f"⏱️ Search timeout exceeded - returning empty result")
        try:
            conn.close()
        except:
            pass
        if hasattr(signal, 'SIGALRM'):
            signal.alarm(0)
        return {
            "products": [],
            "search_method": "timeout",
            "error": "search_timeout"
        }
    except Exception as e:
        print(f"ERROR: Search error: {e}")
        try:
            conn.close()
        except:
            pass
        if hasattr(signal, 'SIGALRM'):
            signal.alarm(0)
        return {
            "products": [],
            "search_method": "error",
            "error": str(e)
        }
    finally:
        # Luôn cancel alarm
        if hasattr(signal, 'SIGALRM'):
            try:
                signal.alarm(0)
            except:
                pass
    
    # Continue với logic cũ nếu có candidates
    try:
        
        print(f"SUCCESS: Found {len(candidates)} candidates with '{main_word}'")
        
        # BƯỚC 2: Tính vector similarity cho từ PHỤ
        # Tăng ngưỡng để loại bỏ sản phẩm không liên quan
        SIMILARITY_THRESHOLD = 0.35  
        MIN_SECONDARY_MATCH_RATIO = 0.5  # Tối thiểu 50% từ phụ phải khớp
        
        # Tạo vector cho query PHỤ (không bao gồm từ chính)
        if secondary_words:
            secondary_query = " ".join(secondary_words)
            secondary_vector = generate_embedding_qwen(secondary_query)
        else:
            # Nếu không có từ phụ, dùng toàn bộ query
            secondary_vector = vector
            # Không cần filter nếu chỉ có 1 từ
            MIN_SECONDARY_MATCH_RATIO = 0
        
        # Tính similarity cho từng candidate
        scored_products = []
        for candidate in candidates:
            product_name = candidate["product_name"].lower()
            
            # Tính vector similarity
            if candidate["description_embedding"] and secondary_vector:
                # Convert embedding từ string hoặc list sang numpy array
                candidate_emb = candidate["description_embedding"]
                if isinstance(candidate_emb, str):
                    candidate_emb = json.loads(candidate_emb)
                
                candidate_np = np.array(candidate_emb)
                query_np = np.array(secondary_vector)
                
                # Cosine similarity
                dot_product = np.dot(candidate_np, query_np)
                norm_a = np.linalg.norm(candidate_np)
                norm_b = np.linalg.norm(query_np)
                similarity = dot_product / (norm_a * norm_b) if (norm_a * norm_b) > 0 else 0
                similarity = float(similarity)
            else:
                similarity = 0.0
            
            # Đếm số từ phụ khớp chính xác
            secondary_match_count = sum(1 for word in secondary_words if word in product_name)
            secondary_match_ratio = secondary_match_count / len(secondary_words) if secondary_words else 1.0
            
            # Tính final score - ƯU TIÊN exact match HƠN
            final_score = (secondary_match_ratio * 0.6) + (similarity * 0.4)
            
            # Thêm vào list scored_products
            scored_products.append({
                "headcode": candidate["headcode"],
                "product_name": candidate["product_name"],
                "category": candidate.get("category"),
                "sub_category": candidate.get("sub_category"),
                "material_primary": candidate.get("material_primary"),
                "project": candidate.get("project"),
                "project_id": candidate.get("project_id"),
                "similarity": round(similarity, 3),
                "secondary_match_count": secondary_match_count,
                "secondary_match_ratio": round(secondary_match_ratio, 2),
                "final_score": round(final_score, 3)
            })
        
        # Lọc theo ĐIỀU KIỆN CHẶT:
        # 1. Similarity >= ngưỡng
        # 2. Nếu có từ phụ: phải khớp tối thiểu 50% từ phụ HOẶC similarity rất cao (>0.6)
        filtered_products = []
        for p in scored_products:
            # Điều kiện 1: Similarity đạt ngưỡng cơ bản
            if p["similarity"] < SIMILARITY_THRESHOLD:
                continue
            
            # Điều kiện 2: Nếu có từ phụ, phải khớp đủ từ hoặc similarity rất cao
            if secondary_words:
                if p["secondary_match_ratio"] >= MIN_SECONDARY_MATCH_RATIO or p["similarity"] >= 0.6:
                    filtered_products.append(p)
            else:
                # Không có từ phụ thì chỉ cần similarity
                filtered_products.append(p)
        
        # Sort theo final_score
        filtered_products.sort(key=lambda x: x["final_score"], reverse=True)
        
        # Giới hạn 10 sản phẩm
        filtered_products = filtered_products[:10]
        
        if filtered_products:
            print(f"SUCCESS: Final: {len(filtered_products)} products (main: '{main_word}', secondary match, similarity >= {SIMILARITY_THRESHOLD})")
            for i, p in enumerate(filtered_products[:3], 1):
                print(f"  {i}. {p['product_name']} (score: {p['final_score']}, sim: {p['similarity']})")
            
            conn.close()
            return {
                "products": filtered_products,
                "search_method": "two_step_main_word_vector",
                "expanded_query": expanded,
                "main_word": main_word,
                "secondary_words": secondary_words
            }
        else:
            print(f"ERROR: No products meet similarity threshold (>= {SIMILARITY_THRESHOLD})")
            conn.close()
            return {
                "products": [],
                "search_method": "no_match_after_filtering",
                "main_word": main_word
            }
            
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
    
    conn.close()
    return {"products": [], "search_method": "hybrid_failed"}

def expand_search_query(user_query: str, params: Dict) -> str:
    """AI mở rộng query ngắn thành mô tả chi tiết với từ khóa chính xác"""
    model = genai.GenerativeModel("gemini-2.5-flash")
    
    prompt = f"""
            Người dùng tìm: "{user_query}"

            Tạo mô tả tìm kiếm tối ưu (2-3 câu ngắn), GIỮ NGUYÊN TỪ KHÓA CHÍNH từ câu gốc:
            1. LOẠI SẢN PHẨM CHÍNH XÁC (bàn/ghế/tủ...) - PHẢI khớp với từ khóa gốc
            2. VẬT LIỆU CỤ THỂ (gỗ teak/đá marble/da bò...)
            3. VỊ TRÍ/CÔNG DỤNG (nhà bếp/phòng khách/dining/coffee...)

            QUAN TRỌNG: 
            - NẾU người dùng tìm "bàn làm việc" thì PHẢI nhấn mạnh "bàn làm việc", "desk", "working table"
            - KHÔNG mở rộng sang loại sản phẩm khác (ví dụ: tìm "bàn" thì không nhắc đến "ghế")
            - Chỉ bổ sung từ đồng nghĩa và chi tiết về loại sản phẩm CỤ THỂ đang tìm

            VD: 
            - "bàn làm việc" -> "Bàn làm việc desk working table văn phòng. Office desk bàn học bàn máy tính."
            - "bàn gỗ teak" -> "Bàn làm từ gỗ teak tự nhiên. Dining table hoặc coffee table chất liệu teak wood cao cấp."

            Output (chỉ mô tả, tập trung vào từ khóa chính):
        """
    
    try:
        response = call_gemini_with_retry(model, prompt, max_retries=2)
        if response:
            # Đảm bảo từ khóa gốc có trong expanded query
            expanded = response.strip()
            if user_query.lower() not in expanded.lower():
                expanded = f"{user_query} {expanded}"
            print(f"Expanded: '{user_query}' -> '{expanded[:100]}...'")
            return expanded
    except:
        pass
    return user_query

def get_latest_material_price(material_subprice_json: str) -> float:
    """Lấy giá mới nhất từ JSON lịch sử giá"""
    if not material_subprice_json:
        return 0.0
    
    try:
        price_history = json.loads(material_subprice_json)
        if not price_history or not isinstance(price_history, list):
            return 0.0
        
        sorted_prices = sorted(
            price_history, 
            key=lambda x: x.get('date', '1900-01-01'), 
            reverse=True
        )
        
        return float(sorted_prices[0].get('price', 0))
    except:
        return 0.0

def search_products_keyword_only(params: Dict):
    """TIER 3: Fallback keyword search"""
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    conditions = []
    values = []
    
    if params.get("category"):
        cat = params['category']
        conditions.append("(category ILIKE %s OR sub_category ILIKE %s OR product_name ILIKE %s)")
        values.extend([f"%{cat}%", f"%{cat}%", f"%{cat}%"])
    
    if params.get("material_primary"):
        mat = params['material_primary']
        conditions.append("(material_primary ILIKE %s OR product_name ILIKE %s)")
        values.extend([f"%{mat}%", f"%{mat}%"])
    
    if conditions:
        where_clause = " OR ".join(conditions)
        sql = f"SELECT * FROM products_qwen WHERE {where_clause} LIMIT 12"
    else:
        sql = "SELECT * FROM products_qwen ORDER BY RANDOM() LIMIT 10"
        values = []
    
    try:
        cur.execute(sql, values)
        results = cur.fetchall()
        conn.close()
        
        if not results:
            return {
                "response": "Không tìm thấy sản phẩm phù hợp.",
                "products": []
            }
        
        print(f"SUCCESS: TIER 3 => Found {len(results)} products")
        products = []
        for r in results:
            product = dict(r)
            product["total_cost"] = calculate_product_total_cost(product["headcode"])
            products.append(product)
        
        return {
            "products": products,
            "search_method": "keyword"
        }
    except Exception as e:
        conn.close()
        print(f"ERROR: TIER 3 failed: {e}")
        return {
            "response": "Lỗi tìm kiếm.",
            "products": []
        }

def calculate_personalized_score(
    candidate_vector: list, 
    session_id: str
) -> float:
    """
    🎯 V5.7 - Trả về điểm Personalization RIÊNG (0.0 → 1.0)
    KHÔNG trả về final_score, để search_products tổng hợp sau
    
    Returns:
        float: Personal affinity score (0.0 = không khớp, 1.0 = rất khớp)
    """
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Lấy 10 interactions gần nhất
        cur.execute("""
            SELECT product_vector, weight
            FROM user_preferences
            WHERE session_id = %s
            ORDER BY created_at DESC
            LIMIT 10
        """, (session_id,))
        
        history = cur.fetchall()
        conn.close()
        
        if not history:
            return 0.5  # Neutral score khi chưa có history
        
        # Convert candidate sang numpy
        if isinstance(candidate_vector, str):
            candidate_np = np.array(json.loads(candidate_vector), dtype=np.float32)
        else:
            candidate_np = np.array(candidate_vector, dtype=np.float32)
        
        positive_scores = []
        negative_scores = []
        
        for record in history:
            try:
                vec_data = record['product_vector']
                
                if vec_data is None:
                    continue
                    
                # Parse vector
                if isinstance(vec_data, str):
                    hist_vector = np.array(json.loads(vec_data), dtype=np.float32)
                elif isinstance(vec_data, list):
                    hist_vector = np.array(vec_data, dtype=np.float32)
                else:
                    continue
                
                # Check dimension match
                if len(hist_vector) != len(candidate_np):
                    continue
                
                # Cosine Similarity
                norm_product = np.linalg.norm(candidate_np) * np.linalg.norm(hist_vector)
                if norm_product < 1e-8:
                    continue
                    
                similarity = np.dot(candidate_np, hist_vector) / norm_product
                
                # Phân loại theo weight
                if record['weight'] > 0:
                    positive_scores.append(similarity)
                else:
                    negative_scores.append(similarity)
                    
            except Exception:
                continue
        
        # Fallback nếu không có scores hợp lệ
        if not positive_scores and not negative_scores:
            return 0.5
        
        # Tính điểm affinity thuần túy
        positive_affinity = np.mean(positive_scores) if positive_scores else 0.0
        negative_penalty = np.mean(negative_scores) if negative_scores else 0.0
        
        # Formula: Positive boost - Negative penalty
        personal_score = positive_affinity - (negative_penalty * 0.5)
        
        # Clip về [0, 1]
        personal_score = float(np.clip(personal_score, 0.0, 1.0))
        
        return personal_score
        
    except Exception as e:
        print(f"WARNING: Personalization error: {e}")
        return 0.5

def generate_consolidated_report(product_headcodes: List[str]) -> BytesIO:
    """
    Tạo báo cáo Excel tổng hợp định mức vật tư cho nhiều sản phẩm
    
    Args:
        product_headcodes: Danh sách mã sản phẩm
    
    Returns:
        BytesIO: File Excel buffer
    """
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    # 1. LẤY THÔNG TIN SẢN PHẨM
    cur.execute("""
        SELECT headcode, product_name, category, sub_category, project
        FROM products_qwen 
        WHERE headcode = ANY(%s)
        ORDER BY product_name
    """, (product_headcodes,))
    
    selected_products = cur.fetchall()
    
    if not selected_products:
        raise ValueError("Không tìm thấy sản phẩm nào")
    
    # 2. LẤY ĐỊNH MỨC CHI TIẾT (Flatten View)
    cur.execute("""
        SELECT 
            p.headcode,
            p.product_name,
            m.id_sap,
            m.material_name,
            m.material_group,
            m.material_subgroup,
            m.unit as material_unit,
            pm.quantity,
            pm.unit as pm_unit,
            m.material_subprice
        FROM product_materials pm
        INNER JOIN products_qwen p ON pm.product_headcode = p.headcode
        INNER JOIN materials m ON pm.material_id_sap = m.id_sap
        WHERE p.headcode = ANY(%s)
        ORDER BY p.product_name, m.material_name
    """, (product_headcodes,))
    
    detail_records = cur.fetchall()
    conn.close()
    
    if not detail_records:
        raise ValueError("Các sản phẩm này chưa có định mức vật tư")
    
    # 3. AGGREGATION - GỘP VẬT TƯ
    material_summary = {}
    
    for record in detail_records:
        id_sap = record['id_sap']
        quantity = float(record['quantity']) if record['quantity'] else 0.0
        
        # Parse giá mới nhất
        latest_price = get_latest_material_price(record['material_subprice'])
        
        if id_sap not in material_summary:
            material_summary[id_sap] = {
                'id_sap': id_sap,
                'material_name': record['material_name'],
                'material_group': record['material_group'],
                'material_subgroup': record['material_subgroup'],
                'unit': record['material_unit'],
                'total_quantity': 0.0,
                'unit_price': latest_price,
                'total_cost': 0.0,
                'used_in_products': []
            }
        
        # Cộng dồn số lượng
        material_summary[id_sap]['total_quantity'] += quantity
        material_summary[id_sap]['used_in_products'].append(
            f"{record['product_name']} ({quantity} {record['pm_unit']})"
        )
    
    # Tính thành tiền
    for mat_id, mat_data in material_summary.items():
        mat_data['total_cost'] = mat_data['total_quantity'] * mat_data['unit_price']
    
    # 4. TẠO EXCEL FILE
    wb = Workbook()
    
    # --- SHEET 1: OVERVIEW (Danh sách SP đã chọn) ---
    ws_overview = wb.active
    ws_overview.title = "Overview"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Headers
    overview_headers = ["STT", "Mã SP", "Tên Sản Phẩm", "Danh Mục", "Dự Án"]
    for col_idx, header in enumerate(overview_headers, 1):
        cell = ws_overview.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for idx, prod in enumerate(selected_products, 1):
        ws_overview.append([
            idx,
            prod['headcode'],
            prod['product_name'],
            f"{prod.get('category', '')} - {prod.get('sub_category', '')}",
            prod.get('project', '')
        ])
    
    # Auto-adjust column width
    for col in ws_overview.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws_overview.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    # --- SHEET 2: MATERIAL SUMMARY (Tổng hợp vật tư) ---
    ws_summary = wb.create_sheet("Material Summary")
    
    summary_headers = [
        "STT", "Mã SAP", "Tên Vật Liệu", "Nhóm", 
        "Nhóm Con", "Đơn Vị", "Tổng SL", "Đơn Giá (VNĐ)", "Thành Tiền (VNĐ)"
    ]
    
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Sort by total_cost DESC
    sorted_materials = sorted(
        material_summary.values(), 
        key=lambda x: x['total_cost'], 
        reverse=True
    )
    
    total_cost_all = 0.0
    
    for idx, mat in enumerate(sorted_materials, 1):
        ws_summary.append([
            idx,
            mat['id_sap'],
            mat['material_name'],
            mat['material_group'],
            mat['material_subgroup'],
            mat['unit'],
            round(mat['total_quantity'], 2),
            round(mat['unit_price'], 2),
            round(mat['total_cost'], 2)
        ])
        total_cost_all += mat['total_cost']
    
    # TỔNG CỘNG ROW
    summary_row = ws_summary.max_row + 1
    ws_summary.cell(row=summary_row, column=7, value="TỔNG CỘNG:").font = Font(bold=True)
    ws_summary.cell(row=summary_row, column=9, value=round(total_cost_all, 2)).font = Font(bold=True, color="FF0000")
    
    for col in ws_summary.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws_summary.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    # --- SHEET 3: DETAILS (Chi tiết theo SP) ---
    ws_details = wb.create_sheet("Details")
    
    detail_headers = [
        "Mã SP", "Tên SP", "Mã SAP", "Tên Vật Liệu", 
        "Nhóm VL", "Số Lượng", "Đơn Vị", "Đơn Giá", "Thành Tiền"
    ]
    
    for col_idx, header in enumerate(detail_headers, 1):
        cell = ws_details.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for record in detail_records:
        quantity = float(record['quantity']) if record['quantity'] else 0.0
        unit_price = get_latest_material_price(record['material_subprice'])
        total_cost = quantity * unit_price
        
        ws_details.append([
            record['headcode'],
            record['product_name'],
            record['id_sap'],
            record['material_name'],
            record['material_group'],
            round(quantity, 2),
            record['pm_unit'],
            round(unit_price, 2),
            round(total_cost, 2)
        ])
    
    for col in ws_details.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws_details.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    # 5. SAVE TO BUFFER
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer
